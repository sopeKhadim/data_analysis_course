# /// script
# requires-python = ">=3.13"
# dependencies = [
#     "marimo>=0.23.9",
#     "setuptools",
#     "ydata-profiling",
# ]
# ///

import marimo

__generated_with = "0.23.9"
app = marimo.App(
    width="medium",
    app_title="Séance 2 — Collecte, Nettoyage et Préparation des données",
)

with app.setup:
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    import seaborn as sns
    import warnings
    warnings.filterwarnings('ignore')
    plt.rcParams['figure.figsize'] = (10, 5)
    plt.rcParams['axes.spines.top'] = False
    plt.rcParams['axes.spines.right'] = False
    sns.set_palette("husl")
    print("✅ Imports OK")
    print(f"   pandas {pd.__version__}  |  numpy {np.__version__}")


@app.cell(hide_code=True)
def imports_marimo():
    import marimo as mo

    return (mo,)


@app.cell(hide_code=True)
def _header(mo):
    mo.md(r"""
    # 📦 Séance 2 — Collecte, Nettoyage, Préparation des données et Echatillonnage.

    **Cours : Analyse de données · 1ère année ingénieur · 2025–2026**

    ---

    > *"Data scientists spend 80% of their time cleaning and preparing data."*
    > — Wes McKinney, *Python for Data Analysis* (O'Reilly, 3e éd., 2022)

    ## Objectifs de la séance

    | Partie | Durée | Contenu |
    |--------|-------|---------|
    | **CM** | 3h00 | Collecte · Valeurs manquantes · Outliers · Normalisation · Profiling · Echantillonage |
    | **TD** | 1h00 | Exercices sur dataset e-commerce |
    | **TP** | 30min | Pipeline de nettoyage complet sur Online Retail |

    ## Dataset utilisé

    | Dataset | Source | Rôle pédagogique |
    |---------|--------|-----------------|
    | **Online Retail** | UCI ML Repository (synthétique) | Fil rouge — exemples pandas & polars dans chaque section + TD et TP complets |

    ## Plan

    1. Collecte des données — CSV, JSON, Excel, API REST, SQL · **🛒 Online Retail pandas vs polars**
    2. Qualité des données — Doublons, cohérences, types · **🛒 Audit manquants Online Retail**
    3. Valeurs manquantes — MCAR / MAR / MNAR · Stratégies · **🛒 Imputation UnitPrice, CustomerID**
    4. Détection des outliers — IQR · Z-score · Isolation Forest
    5. Normalisation et standardisation — Min-Max · Z-score · Robust · **🛒 Quantity et UnitPrice**
    6. Encodage des variables catégorielles — Label · OneHot · Target · **🛒 Country, Description** *(à venir)*
    7. Profiling automatique — ydata-profiling
    8. Feature Engineering — Extraction dates · Binning · Combinaison *(à venir)*
    9. Pipeline complet — ColumnTransformer
    10. Échantillonnage — SRS · Stratifié · Grappes · Bootstrap · Biais · Classes déséquilibrées · SMOTE
    11. TD — Exercices e-commerce
    12. TP — Pipeline de nettoyage complet

    > **Note :** Ce cours fait suite à la Séance 1. Les concepts de `.head()`, `.info()`, `.describe()`, et la manipulation de base avec pandas sont considérés comme acquis. La Séance 1 a introduit `SimpleImputer`, `OrdinalEncoder`, `OneHotEncoder` et les pipelines scikit-learn. Nous allons ici **approfondir** avec des méthodes plus robustes (MCAR/MAR/MNAR, KNN, Isolation Forest, RobustScaler) et introduire **polars** comme alternative moderne à pandas.
    """)
    return


@app.cell(hide_code=True)
def _section1(mo):
    mo.md(r"""
    ---
    ## 1. Collecte des données

    La première étape de tout projet data est **d'acquérir les données** depuis leurs sources.

    | Source | Format typique | Outil Python |
    |--------|---------------|--------------|
    | Fichiers plats | CSV, TSV | `pd.read_csv()` |
    | Excel | XLSX, XLS | `pd.read_excel()` |
    | JSON / APIs | JSON | `pd.read_json()`, `requests` |
    | Bases de données SQL | SQL | `pd.read_sql()`, SQLAlchemy |
    | Parquet / Arrow | Colonnes compressées | `pd.read_parquet()` |

    ### 1.1 Chargement CSV — Options avancées

    ```python
    df = pd.read_csv(
    "data.csv",
    sep=";",               # séparateur point-virgule (format FR)
    decimal=",",           # virgule décimale
    encoding="latin-1",    # encodage Windows
    parse_dates=["date"],  # conversion datetime automatique
    na_values=["N/A", "?", "-"],  # valeurs → NaN
    dtype={"code_postal": str},    # forcer le type
    )
    ```

    ### 1.2 Collecte via API REST

    ```python
    import requests
    import pandas as pd

    url = "https://api.open-meteo.com/v1/forecast"
    params = {
    "latitude": 14.6928,   # Dakar
    "longitude": -17.4467,
    "daily": "temperature_2m_max,temperature_2m_min",
    "forecast_days": 7,
    }
    response = requests.get(url, params=params)

    if response.status_code == 200:
    data = response.json()
    df_meteo = pd.DataFrame(data["daily"])
    ```

    **Codes HTTP importants :** `200` OK · `404` Non trouvé · `429` Trop de requêtes · `403` Interdit

    > ⚠️ **Sécurité :** Stocker les clés API dans des variables d'environnement (`os.environ`), jamais dans le code.

    ### 1.3 Collecte depuis une base SQL

    ```python
    import sqlalchemy as sa
    import pandas as pd

    engine = sa.create_engine("postgresql://user:password@localhost:5432/mydb")

    # Requête paramétrée (évite les injections SQL)
    query = sa.text("SELECT * FROM commandes WHERE region = :region")
    df = pd.read_sql(query, engine, params={"region": "Dakar"})
    ```
    """)
    return


@app.cell(hide_code=True)
def _retail_collecte_theory(mo):
    mo.md(r"""
    ### 🛒 Exemple fil rouge — Online Retail (pandas vs polars)

    Le dataset **Online Retail** (UCI ML Repository) est notre référence tout au long du cours.
    Il illustre chaque concept dans un contexte métier concret : transactions e-commerce d'un grossiste britannique
    de décembre 2010 à décembre 2011 (~541 909 lignes, 8 colonnes).

    **pandas** est la bibliothèque historique, **polars** est une alternative moderne (Rust, lazy evaluation, API expressif).

    | | **pandas** | **polars** |
    |---|---|---|
    | Langage interne | Python/NumPy | Rust |
    | Évaluation | Eager (immédiate) | Lazy ou Eager |
    | Syntaxe sélection | `df["col"]` / `.loc` | `df.select(pl.col("col"))` |
    | Chaînage | Limité | Natif (`.pipe`, `.lazy()`) |
    | Vitesse | Référence | 5–20× plus rapide sur gros volumes |

    ```python
    # pandas — options adaptées au fichier Online Retail (Excel d'origine)
    import pandas as pd

    df = pd.read_excel("Online Retail.xlsx",
                       dtype={"CustomerID": str},
                       parse_dates=["InvoiceDate"])

    # Ou depuis un CSV exporté
    df = pd.read_csv("online_retail.csv",
                     encoding="latin-1",        # encodage Windows fréquent
                     parse_dates=["InvoiceDate"],
                     dtype={"CustomerID": str},
                     na_values=["\", "NA", "?"])

    # polars — chargement direct, très rapide sur ~500k lignes
    import polars as pl
    df_pl = pl.read_csv("online_retail.csv",
                        null_values=["\", "NA", "?"],
                        try_parse_dates=True)
    ```
    """)
    return


@app.cell
def _retail_chargement_demo():
    """Démonstration du chargement — utilise le dataset synthétique créé en mémoire."""
    import io as _io

    # Simulation d'un chargement CSV avec options avancées (identique à un vrai fichier Online Retail)
    print("=== 🛒 ONLINE RETAIL — Chargement pandas (options avancées) ===")
    print()
    print("Exemple de commande réelle pour le fichier UCI :")
    print("""  df = pd.read_csv(
      'online_retail.csv',
      encoding='latin-1',          # encodage Windows fréquent
      parse_dates=['InvoiceDate'],
      dtype={'CustomerID': str},   # éviter la conversion float → perte du zéro de tête
      na_values=['', 'NA', '?'],
      )""")
    print()
    print("=== Dataset synthétique chargé en mémoire (500 lignes, 8 colonnes) ===")
    print("Colonnes : InvoiceNo, StockCode, Description, Quantity, InvoiceDate, UnitPrice, CustomerID, Country")
    print()
    print("Types attendus après chargement :")
    _schema = {
        "InvoiceNo"  : "object  — numéro de facture (peut contenir 'C' pour retour)",
        "StockCode"  : "object  — code article (alphanumérique)",
        "Description": "object  — libellé produit (peut être NaN)",
        "Quantity"   : "int64   — quantité (négatif = retour/annulation)",
        "InvoiceDate": "datetime64 — horodatage de la transaction",
        "UnitPrice"  : "float64 — prix unitaire en £ (peut être NaN ou 0)",
        "CustomerID" : "object  — identifiant client (~20% manquant = guest checkout)",
        "Country"    : "object  — pays de livraison (65% UK)",
    }
    for _col, _desc in _schema.items():
        print(f"  {_col:<14} → {_desc}")

    # Polars (si installé)
    print()
    try:
        import polars as _pl
        import pandas as _pd_demo
        import numpy as _np_demo

        _n_demo = 10
        _df_demo = _pd_demo.DataFrame({
            "InvoiceNo"  : [f"INV{i:05d}" for i in range(_n_demo)],
            "StockCode"  : ["85123A", "71053", "84406B", "84029G", "84029E",
                            "22752", "21730", "85123A", None, "71053"],
            "Description": ["WHITE HANGING HEART T-LIGHT HOLDER", "WHITE METAL LANTERN",
                             None, "KNITTED UNION FLAG HOT WATER BOTTLE", "RED WOOLLY HOTTIE",
                             "CREAM CUPID HEARTS COAT HANGER", "GLASS STAR FROSTED T-LIGHT HOLDER",
                             "WHITE HANGING HEART T-LIGHT HOLDER", "SAMPLE", None],
            "Quantity"   : [6, 6, 8, 6, 6, 2, 6, -1, 32, 6],
            "UnitPrice"  : [2.55, 3.39, 2.75, 3.39, 3.39, None, 3.25, 2.55, 5.95, 3.39],
            "CustomerID" : ["17850", "13047", None, "13047", "17850",
                             "17850", None, "17850", "13047", None],
            "Country"    : ["United Kingdom", "United Kingdom", "France", "United Kingdom", "United Kingdom",
                             "Germany", "United Kingdom", "United Kingdom", None, "France"],
        })
        _df_pl_demo = _pl.from_pandas(_df_demo)
        print("=== Polars — schema() et head() ===")
        print(_df_pl_demo.schema)
        print(_df_pl_demo.head(5))
    except ImportError:
        print("💡 polars non installé (uv add polars) — exemples polars affichés en commentaire")
    return


@app.cell
def _create_dataset():
    """
    Dataset synthétique Online Retail-like avec imperfections réalistes.
    Source d'inspiration : Online Retail Dataset (UCI ML Repository)
    https://archive.ics.uci.edu/dataset/352/online+retail
    """
    np.random.seed(42)
    _n = 500

    _invoice_no = [f"INV{str(i).zfill(5)}" for i in range(1, _n + 1)]
    _stock_codes = np.random.choice(
        ["85123A", "71053", "84406B", "84029G", "84029E", "22752", "21730", None],
        _n, p=[0.15, 0.15, 0.15, 0.15, 0.15, 0.10, 0.10, 0.05]
    )
    _descriptions = np.random.choice(
        ["WHITE HANGING HEART T-LIGHT HOLDER", "WHITE METAL LANTERN",
         "CREAM CUPID HEARTS COAT HANGER", "KNITTED UNION FLAG HOT WATER BOTTLE",
         "RED WOOLLY HOTTIE WHITE HEART", None, "???", "SAMPLE"],
        _n, p=[0.20, 0.20, 0.15, 0.15, 0.15, 0.05, 0.05, 0.05]
    )
    _quantities = np.concatenate([
        np.random.randint(1, 100, int(_n * 0.85)),
        np.random.randint(-50, 0, int(_n * 0.05)),
        np.array([500, 600, 800, 1000, 2000]),
        np.zeros(_n - int(_n * 0.85) - int(_n * 0.05) - 5, dtype=int)
    ])[:_n]

    _unit_prices = np.abs(np.random.exponential(5, _n))
    _unit_prices[np.random.choice(_n, 30, replace=False)] = np.nan
    _unit_prices = np.concatenate([_unit_prices[:495], [0.0, 0.0, 150.0, 200.0, 500.0]])[:_n]

    _customer_ids = np.where(
        np.random.random(_n) > 0.20,
        np.random.choice(range(12000, 18000), _n),
        None
    )
    _countries = np.random.choice(
        ["United Kingdom", "France", "Germany", "Spain", "Netherlands", "EIRE", None],
        _n, p=[0.65, 0.10, 0.08, 0.07, 0.05, 0.04, 0.01]
    )
    _dates = pd.date_range("2010-12-01", "2011-12-09", periods=_n)

    df_retail = pd.DataFrame({
        "InvoiceNo": _invoice_no,
        "StockCode": _stock_codes,
        "Description": _descriptions,
        "Quantity": _quantities,
        "InvoiceDate": _dates,
        "UnitPrice": _unit_prices,
        "CustomerID": _customer_ids,
        "Country": _countries,
    })

    # Ajout de doublons intentionnels
    _doublons = df_retail.sample(15, random_state=1)
    df_retail = pd.concat([df_retail, _doublons], ignore_index=True).sample(frac=1, random_state=42).reset_index(drop=True)

    print(f"✅ Dataset créé : {df_retail.shape[0]} lignes × {df_retail.shape[1]} colonnes")
    print(df_retail.head(5).to_string())
    return (df_retail,)


@app.cell(hide_code=True)
def _section2(mo):
    mo.md(r"""
    ---
    ## 2. Qualité des données

    Avant de traiter les données, il faut **diagnostiquer leur état de santé**.

    | Problème | Description | Exemple |
    |----------|-------------|---------|
    | **Doublons** | Lignes identiques | Même commande enregistrée deux fois |
    | **Valeurs manquantes** | `NaN`, `None` | Client sans identifiant |
    | **Types incohérents** | Nombre stocké en texte | `"123"` au lieu de `123` |
    | **Valeurs aberrantes** | Hors plage normale | Prix = -5€ |
    | **Incohérences métier** | Logique invalide | Date fin < Date début |
    """)
    return


@app.cell
def _audit(df_retail):
    print("=" * 60)
    print("AUDIT RAPIDE DU DATASET")
    print("=" * 60)
    print(f"\n📐 Dimensions : {df_retail.shape[0]:,} lignes × {df_retail.shape[1]} colonnes")
    print("\n📋 Types de données :")
    print(df_retail.dtypes.to_string())
    _missing = df_retail.isnull().sum()
    _missing_pct = (_missing / len(df_retail) * 100).round(1)
    _missing_df = pd.DataFrame({"Manquants": _missing, "%": _missing_pct})
    print("\n🔍 Valeurs manquantes :")
    print(_missing_df[_missing_df["Manquants"] > 0].to_string())
    print(f"\n🔁 Doublons : {df_retail.duplicated().sum()} lignes")
    print("\n📊 Statistiques (numériques) :")
    print(df_retail.describe().round(2).to_string())
    return


@app.cell
def _plot_manquants(df_retail):
    import seaborn as _sns2
    _fig1, _axes1 = plt.subplots(1, 2, figsize=(14, 5))

    _mp = (df_retail.isnull().sum() / len(df_retail) * 100).sort_values(ascending=True)
    _mp = _mp[_mp > 0]
    _colors1 = ["#e74c3c" if p > 30 else "#f39c12" if p > 10 else "#3498db" for p in _mp]
    _mp.plot(kind="barh", ax=_axes1[0], color=_colors1)
    _axes1[0].set_xlabel("% de valeurs manquantes")
    _axes1[0].set_title("Valeurs manquantes par colonne", fontweight="bold")
    _axes1[0].axvline(x=30, color="red", linestyle="--", alpha=0.5, label="Seuil 30%")
    _axes1[0].axvline(x=10, color="orange", linestyle="--", alpha=0.5, label="Seuil 10%")
    _axes1[0].legend(fontsize=9)
    for _i1, (_idx1, _v1) in enumerate(_mp.items()):
        _axes1[0].text(_v1 + 0.3, _i1, f"{_v1:.1f}%", va="center", fontsize=9)

    _sample1 = df_retail.head(100).isnull().astype(int)
    _sns2.heatmap(_sample1.T, ax=_axes1[1], cmap="Blues", cbar=False,
                  yticklabels=_sample1.columns, xticklabels=False)
    _axes1[1].set_title("Pattern valeurs manquantes\n(100 premières lignes)", fontweight="bold")

    plt.tight_layout()
    plt.suptitle("Diagnostic des valeurs manquantes", y=1.02, fontsize=13, fontweight="bold")
    plt.savefig("/tmp/manquants.png", dpi=120, bbox_inches="tight")
    plt.show()
    print("Rouge = critique (>30%) | Orange = modéré (>10%) | Bleu = faible")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    [**missingno**](https://github.com/residentmario/missingno) est une librairie Python qui vous permet de d'analyser rapidement les valeurs manquantes du dataset. Elle permet d'analyser la corrélation entre les colonnes manquantes
    """)
    return


@app.cell
def _missingno_viz(df_retail):
    try:
        import missingno as msno
        _fig_msno, _axes_msno = plt.subplots(1, 3, figsize=(18, 5))

        # Matrix — zones contiguës = manquants non aléatoires
        msno.matrix(df_retail, ax=_axes_msno[0], sparkline=False, color=(0.2, 0.5, 0.8))
        _axes_msno[0].set_title("Matrix — zones contiguës\n→ manquants non aléatoires", fontweight="bold")

        # Heatmap — corrélations entre colonnes manquantes
        msno.heatmap(df_retail, ax=_axes_msno[1], cmap="RdYlGn")
        _axes_msno[1].set_title("Heatmap — corrélations\nentre colonnes manquantes", fontweight="bold")

        # Dendrogram — regroupements de colonnes
        msno.dendrogram(df_retail, ax=_axes_msno[2], orientation="top")
        _axes_msno[2].set_title("Dendrogramme — colonnes\nà manquants similaires", fontweight="bold")

        plt.suptitle("Analyse structurelle des données manquantes — missingno", fontsize=13, fontweight="bold")
        plt.tight_layout()
        plt.savefig("/tmp/missingno.png", dpi=120, bbox_inches="tight")
        plt.show()
        print("💡 Heatmap : corrélation ≈ +1 → manquants ensemble | ≈ -1 → l'un présent implique l'autre absent") #  corrélations dans les endroits où les données sont manquantes
        print("💡 Dendrogramme : colonnes proches = patterns de manquants similaires → possible MAR") # Les feuilles qui sont au même niveau prédisent la présence de l’autre (vide ou pleine). Les bras verticaux sont utilisés pour indiquer les différents groupes. Les bras courts signifient que les branches sont similaires
        print("")
    except ImportError:
        print("⚠️  missingno non installé — lancez : uv add missingno")
        print("   La librairie génère des visualisations matricielles des patterns de données manquantes")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Suppression des doublons.
    """)
    return


@app.cell
def _doublons(df_retail):
    print("=== GESTION DES DOUBLONS ===")
    print(f"Taille initiale : {len(df_retail):,}")
    print(f"Doublons : {df_retail.duplicated().sum()}")
    df_clean = df_retail.drop_duplicates()
    print(f"Après suppression : {len(df_clean):,} lignes")
    return (df_clean,)


@app.cell(hide_code=True)
def _section3(mo):
    mo.md(r"""
    ---
    ## 3. Valeurs manquantes — MCAR / MAR / MNAR

    Toutes les valeurs manquantes ne sont pas équivalentes. Rubin (1976) distingue trois mécanismes fondamentaux :

    ### 3.1 Taxonomie

    #### MCAR — Missing Completely At Random

    La probabilité qu'une valeur soit manquante est **indépendante** de toutes les données :

    $$P(\text{manquant} \mid X_{\text{obs}}, X_{\text{manq}}) = P(\text{manquant})$$

    **Exemple :** Panne aléatoire d'un capteur de température. Aucun biais introduit.
    **Stratégie :** Imputation simple (médiane, moyenne) acceptable.

    #### MAR — Missing At Random

    La probabilité de manquant **dépend des données observées**, pas des valeurs manquantes :

    $$P(\text{manquant} \mid X_{\text{obs}}, X_{\text{manq}}) = P(\text{manquant} \mid X_{\text{obs}})$$

    **Exemple :** Les hommes répondent moins aux questions sur leur poids (dépend du sexe — observé). Les clients `guest checkout` n'ont pas de `CustomerID` (dépend du mode d'achat — observable).
    **Stratégie :** KNN Imputer, IterativeImputer (MICE), ou indicatrice.

    #### MNAR — Missing Not At Random

    La probabilité de manquant **dépend de la valeur manquante elle-même** :

    $$P(\text{manquant} \mid X_{\text{obs}}, X_{\text{manq}}) = f(X_{\text{manq}})$$

    **Exemple :** Les patients très malades ne remplissent pas le questionnaire de santé — les cas les plus graves sont sous-représentés.
    **Stratégie :** Indicatrice de manquant + modèle de sélection. Expertise métier indispensable.

    > ⚠️ **Le MNAR ne peut pas être détecté uniquement à partir des données observées.** Il nécessite une connaissance du processus de collecte.
    """)
    return


@app.cell
def _retail_audit_manquants(df_retail):
    print("=== 🛒 ONLINE RETAIL — Audit des valeurs manquantes ===\n")
    _df_r = df_retail.copy()

    _mp_r = _df_r.isnull().sum()
    _pct_r = (_mp_r / len(_df_r) * 100).round(1)
    _diag = pd.DataFrame({"Manquants": _mp_r, "%": _pct_r}).sort_values("%", ascending=False)
    print(_diag.to_string())
    print()

    # Classification MCAR / MAR / MNAR
    _analyse = {
        "CustomerID" : ("~20%", "MAR",  "Guest checkout — dépend du mode d'achat (observable)"),
        "UnitPrice"  : ("~6%",  "MCAR", "Panne système aléatoire — indépendant des autres variables"),
        "Description": ("~5%",  "MCAR", "Libellé absent aléatoirement (problème export)"),
        "StockCode"  : ("~5%",  "MCAR", "Code article manquant aléatoirement"),
        "Country"    : ("~1%",  "MCAR", "Pays non renseigné (faible taux, négligeable)"),
    }
    print(f"{'Colonne':<14} {'% manq.':<10} {'Mécanisme':<10} Interprétation")
    print("─" * 78)
    for _col_r, (_pct, _mec, _expl) in _analyse.items():
        print(f"{_col_r:<14} {_pct:<10} {_mec:<10} {_expl}")

    # Visualisation
    _fig_r, _axes_r = plt.subplots(1, 2, figsize=(14, 4))
    _mp_nz = _pct_r[_pct_r > 0].sort_values()
    _colors_r = ["#e74c3c" if p > 30 else "#f39c12" if p > 10 else "#3498db" for p in _mp_nz]
    _mp_nz.plot(kind="barh", ax=_axes_r[0], color=_colors_r)
    _axes_r[0].axvline(30, color="red", linestyle="--", alpha=0.5, label="Seuil 30%")
    _axes_r[0].axvline(10, color="orange", linestyle="--", alpha=0.5, label="Seuil 10%")
    _axes_r[0].set_title("Online Retail — % valeurs manquantes", fontweight="bold")
    _axes_r[0].set_xlabel("% manquants")
    _axes_r[0].legend(fontsize=8)
    for _i_r, (_idx_r, _v_r) in enumerate(_mp_nz.items()):
        _axes_r[0].text(_v_r + 0.3, _i_r, f"{_v_r:.1f}%", va="center", fontsize=9)

    # CustomerID manquant par pays → preuve MAR
    _top5_r = _df_r["Country"].value_counts().head(5).index
    _sub5 = _df_r[_df_r["Country"].isin(_top5_r)]
    _pct_cid = _sub5.groupby("Country")["CustomerID"].apply(lambda x: x.isnull().mean() * 100).sort_values()
    _colors_cid = ["#2ecc71" if v < 15 else "#f39c12" if v < 25 else "#e74c3c" for v in _pct_cid]
    _axes_r[1].barh(_pct_cid.index, _pct_cid.values, color=_colors_cid, alpha=0.8, edgecolor="white")
    _axes_r[1].set_title("CustomerID manquant par pays — preuve MAR\n(dépend du marché → observable)", fontweight="bold")
    _axes_r[1].set_xlabel("% CustomerID manquant")
    for _i_r2, _v_r2 in enumerate(_pct_cid.values):
        _axes_r[1].text(_v_r2 + 0.3, _i_r2, f"{_v_r2:.1f}%", va="center", fontsize=9)
    _note_r = "Taux varie selon le pays\n→ dépend d'une variable observée\n→ mécanisme MAR confirmé"
    _axes_r[1].text(0.97, 0.97, _note_r, transform=_axes_r[1].transAxes,
                    va="top", ha="right", fontsize=9, style="italic",
                    bbox=dict(boxstyle="round", facecolor="#fef9e7", alpha=0.9))

    plt.suptitle("🛒 Online Retail — Diagnostic et mécanismes de données manquantes", fontsize=12, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/retail_manquants.png", dpi=120, bbox_inches="tight")
    plt.show()

    # Polars (si installé)
    try:
        import polars as _pl
        _df_pl_r = _pl.from_pandas(_df_r)
        print("\n📊 Polars — null_count() :")
        print(_df_pl_r.null_count())
    except ImportError:
        print("\n# Polars équivalent :")
        print("# df_pl.null_count()")
        print("# df_pl.select(pl.all().is_null().sum())")
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### Illustration des concepts MCAR, MAR et MNAR dans le dataset ci-dessous
    """)
    return


@app.cell
def _plot_mcar_mar_mnar():
    np.random.seed(0)
    _n2 = 200
    _age2 = np.random.normal(40, 12, _n2).clip(18, 80)
    _salaire2 = 1500 + 60 * _age2 + np.random.normal(0, 500, _n2)

    _fig2, _axes2 = plt.subplots(1, 3, figsize=(15, 5))
    _titres2 = ["MCAR\n(aléatoire complet)", "MAR\n(aléatoire conditionnel)", "MNAR\n(non aléatoire)"]
    _subtitres2 = [
        "Manquant indépendant de tout",
        "Manquant si âge < 30 (observé)",
        "Manquant si salaire élevé (manquant)"
    ]
    _mask_mcar2 = np.random.random(_n2) < 0.35
    _mask_mar2 = _age2 < 30
    _mask_mnar2 = _salaire2 > np.percentile(_salaire2, 65)

    for _ax2, _mask2, _titre2, _sous2 in zip(_axes2, [_mask_mcar2, _mask_mar2, _mask_mnar2], _titres2, _subtitres2):
        _ax2.scatter(_age2[~_mask2], _salaire2[~_mask2], alpha=0.6, s=25, color="#3498db", label="Observé")
        _ax2.scatter(_age2[_mask2], _salaire2[_mask2], alpha=0.6, s=25, color="#e74c3c", marker="x", label="Manquant")
        _ax2.set_xlabel("Âge (observé)")
        _ax2.set_ylabel("Salaire")
        _ax2.set_title(f"{_titre2}", fontweight="bold", fontsize=11)
        _ax2.text(0.5, -0.18, _sous2, transform=_ax2.transAxes, ha="center",
                  fontsize=9, color="gray", style="italic")
        _ax2.legend(fontsize=8)
        _n_m2 = _mask2.sum()
        _ax2.text(0.02, 0.97, f"Manquants : {_n_m2} ({_n_m2/_n2*100:.0f}%)",
                  transform=_ax2.transAxes, va="top", fontsize=9,
                  bbox=dict(boxstyle="round", facecolor="white", alpha=0.8))

    plt.suptitle("Trois mécanismes de données manquantes (Rubin, 1976)", fontsize=13, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig("/tmp/mcar_mar_mnar.png", dpi=120, bbox_inches="tight")
    plt.show()
    return


@app.cell(hide_code=True)
def _strategies_imputation(mo):
    mo.md(r"""
    ### 3.2 Stratégies d'imputation

    | Stratégie | Mécanisme adapté | Avantages | Inconvénients |
    |-----------|-----------------|-----------|---------------|
    | **Suppression** (listwise) | MCAR uniquement | Simple | Perte d'information |
    | **Médiane / Moyenne** | MCAR, MAR simple | Rapide | Réduit la variance |
    | **Mode** (catégoriel) | MCAR | Simple | Peut déséquilibrer |
    | **Forward/Backward fill** | Séries temporelles | Préserve la tendance | Mauvais sur longs gaps |
    | **KNN Imputer** | MAR | Utilise les voisins | Lent (grands datasets) |
    | **IterativeImputer** (MICE) | MAR | Modélise les relations | Complexe |
    | **Indicatrice de manquant** | MNAR | Capture le signal | Combinaison nécessaire |

    **Formule de l'imputation par la médiane :**

    $$x_{\text{imputé}} = \text{médiane}(x_1, x_2, \ldots, x_n)$$

    La médiane est préférée à la moyenne car elle est **robuste aux outliers**.
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ### sklearn.impute
    Le module  [**impute**](https://scikit-learn.org/stable/modules/impute.html) de sklearn offre un ensemble de fonctionnalités qui permet de traiter les valeurs manquantes : ***SimpleImputer***, ***KNNImputer***, ***IterativeImputer***.

     ***SimpleImputer*** : Cette classe applique une stratégie simple pour remplacer les valeurs manquantes par une statistique calculée sur chaque colonne (*mean*, *median*, *most_frequent*). Si la stratégie est *constant*, il faut ajouter le paramètre fill_value avec la valeur souhaitée.
    **Paramètres clés** : *strategy* (strategie de remplacement)

    ***KNNImputer*** : Cette classe impute les valeurs manquantes à partir des k plus proches voisins. Pour chaque valeur manquante, elle calcule la moyenne (simple ou pondérée par la distance) des valeurs observées chez les *n_neighbors* échantillons les plus proches.
    **Paramètres clés** : *n_neighbors* (nombre de voisins, défaut 5) et *weights* (uniform ou distance).

    ***IterativeImputer*** : Cette classe  modélise chaque colonne avec valeurs manquantes comme une fonction des autres colonnes, via un modèle de régression (par défaut BayesianRidge), de manière itérative jusqu'à convergence.
    **Paramètres clés** : *estimator* (modèle de régression), *max_iter* (nombre d'itérations, défaut 10), *initial_strategy* (stratégie d'amorçage).
    """)
    return


@app.cell
def _demo_imputation(df_clean):
    from sklearn.impute import SimpleImputer, KNNImputer
    from sklearn.experimental import enable_iterative_imputer  # noqa
    from sklearn.impute import IterativeImputer

    print("=== STRATÉGIES D'IMPUTATION ===\n")
    _prix3 = df_clean["UnitPrice"].copy() 
    print(f"UnitPrice — manquants : {_prix3.isnull().sum()} ({_prix3.isnull().sum()/len(_prix3)*100:.1f}%)")

    # 1. Médiane
    _imp_med = SimpleImputer(strategy="median")
    _imp_med.fit_transform(_prix3.values.reshape(-1, 1))
    print(f"\n1. Médiane  → valeur imputée = {_imp_med.statistics_[0]:.4f} €")

    # 2. Moyenne
    _imp_mean = SimpleImputer(strategy="mean")
    _imp_mean.fit_transform(_prix3.values.reshape(-1, 1))
    print(f"2. Moyenne  → valeur imputée = {_imp_mean.statistics_[0]:.4f} €")

    # 3. KNN (multi-colonnes pour illustration)
    _X3 = df_clean[["Quantity", "UnitPrice"]].copy().head(200)
    _imp_knn = KNNImputer(n_neighbors=5)
    _X3_knn = _imp_knn.fit_transform(_X3)
    print(f"3. KNN(k=5) → imputation basée sur les 5 voisins les plus proches")

    # 4. MICE (IterativeImputer)
    _imp_iter = IterativeImputer(max_iter=10, random_state=42)
    _X3_mice = _imp_iter.fit_transform(_X3)
    print(f"4. MICE     → modélisation itérative des relations entre variables")

    print("\n📌 Toujours fitter sur le TRAIN SET uniquement, jamais sur le test set !")

    imp_median_val = _imp_med.statistics_[0]
    return (imp_median_val,)


@app.cell
def _plot_imputation(df_clean, imp_median_val):
    _prix_obs4 = df_clean["UnitPrice"]
    _prix_med4 = _prix_obs4.fillna(imp_median_val)
    _prix_mean4 = _prix_obs4.fillna(_prix_obs4.mean())

    _fig4, _axes4 = plt.subplots(1, 3, figsize=(15, 4))
    _bins4 = np.linspace(0, 30, 40)

    for _ax4, _d4, _lbl4, _col4 in zip(
        _axes4,
        [_prix_obs4, _prix_med4, _prix_mean4],
        ["Originale", "Imputation médiane", "Imputation moyenne"],
        ["#95a5a6", "#2ecc71", "#e67e22"]
    ):
        _ax4.hist(_d4[_d4 <= 30], bins=_bins4, color=_col4, alpha=0.7, edgecolor="white")
        _ax4.axvline(_d4.median(), color="red", linestyle="--", label=f"Médiane={_d4.median():.2f}")
        _ax4.axvline(_d4.mean(), color="purple", linestyle=":", label=f"Moyenne={_d4.mean():.2f}")
        _ax4.set_title(_lbl4, fontweight="bold")
        _ax4.legend(fontsize=8)

    plt.suptitle("Impact de la stratégie d'imputation — UnitPrice", fontsize=12, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/imputation.png", dpi=120, bbox_inches="tight")
    plt.show()
    return


@app.cell
def _retail_imputation(df_clean):
    from sklearn.impute import SimpleImputer as _SI_t, KNNImputer as _KNN_t

    print("=== 🛒 ONLINE RETAIL — Stratégies d'imputation par colonne ===\n")
    _df_ri = df_clean[["Quantity", "UnitPrice", "CustomerID", "Country"]].copy()

    print(f"Manquants initiaux :")
    print(f"  UnitPrice  = {_df_ri['UnitPrice'].isnull().sum()} ({_df_ri['UnitPrice'].isnull().mean()*100:.1f}%) → MCAR → médiane")
    print(f"  CustomerID = {_df_ri['CustomerID'].isnull().sum()} ({_df_ri['CustomerID'].isnull().mean()*100:.1f}%) → MAR  → indicatrice")
    print(f"  Country = {_df_ri['Country'].isnull().sum()} ({_df_ri['Country'].isnull().mean()*100:.1f}%) → MCAR → mode")
    print()

    # UnitPrice — MCAR → médiane (simple, non biaisée pour MCAR)
    _prix_avant = _df_ri["UnitPrice"].copy()
    _imp_med_r = _SI_t(strategy="median")
    _df_ri["UnitPrice"] = _imp_med_r.fit_transform(_df_ri[["UnitPrice"]]).flatten()
    print(f"UnitPrice — Médiane : valeur imputée = {_imp_med_r.statistics_[0]:.4f} £")

    # UnitPrice — KNN pour comparaison (utilise Quantity comme voisin)
    _X_knn_r = _df_ri[["Quantity", "UnitPrice"]].copy()
    _X_knn_r["UnitPrice"] = _prix_avant  # restaurer les manquants
    _knn_r = _KNN_t(n_neighbors=5)
    _prix_knn = _knn_r.fit_transform(_X_knn_r)[:, 1]
    print(f"UnitPrice — KNN(k=5) : médiane avant={_prix_avant.median():.4f} | médiane après={pd.Series(_prix_knn).median():.4f}")

    #CustomerID — MAR → indicatrice (ne pas imputer une ID fictive)
    _df_ri["HasCustomerID"] = _df_ri["CustomerID"].notna().astype(int)
    print(f"\nCustomerID — Indicatrice créée : HasCustomerID")
    print(f"  Clients identifiés : {_df_ri['HasCustomerID'].sum()} ({_df_ri['HasCustomerID'].mean()*100:.1f}%)")
    print(f"  Guest checkout     : {(1-_df_ri['HasCustomerID']).sum()} ({(1-_df_ri['HasCustomerID']).mean()*100:.1f}%)")

    # Country — MCAR → mode
    _mode_country = _df_ri["Country"].mode()[0]
    _df_ri["Country"] = _df_ri["Country"].fillna(_mode_country)
    print(f"\nCountry — Mode : valeur imputée = '{_mode_country}' (MCAR)")
    print(f"\nManquants après imputation : {_df_ri.isnull().sum().sum()} (hors CustomerID conservé comme objet)")

    # Visualisation : Médiane vs KNN pour UnitPrice
    _fig_ri, _axes_ri = plt.subplots(1, 3, figsize=(15, 4))
    _bins_ri = np.linspace(0, 20, 40)

    _axes_ri[0].hist(_prix_avant.dropna()[_prix_avant.dropna() <= 20], bins=_bins_ri,
                     color="#95a5a6", alpha=0.8, edgecolor="white")
    _axes_ri[0].axvline(_prix_avant.median(), color="red", linestyle="--",
                        label=f"Médiane={_prix_avant.median():.2f} £")
    _axes_ri[0].set_title("UnitPrice — original\n(avec manquants)", fontweight="bold")
    _axes_ri[0].set_xlabel("UnitPrice (£)")
    _axes_ri[0].legend(fontsize=8)

    _prix_med_full = _prix_avant.fillna(_prix_avant.median())
    _axes_ri[1].hist(_prix_med_full[_prix_med_full <= 20], bins=_bins_ri,
                     color="#e67e22", alpha=0.8, edgecolor="white")
    _axes_ri[1].axvline(_prix_med_full.median(), color="red", linestyle="--",
                        label=f"Médiane={_prix_med_full.median():.2f} £")
    _axes_ri[1].set_title("UnitPrice — imputation médiane\n(pic artificiel à la médiane)", fontweight="bold")
    _axes_ri[1].set_xlabel("UnitPrice (£)")
    _axes_ri[1].legend(fontsize=8)

    _prix_knn_s = pd.Series(_prix_knn)
    _axes_ri[2].hist(_prix_knn_s[_prix_knn_s <= 20], bins=_bins_ri,
                     color="#2ecc71", alpha=0.8, edgecolor="white")
    _axes_ri[2].axvline(_prix_knn_s.median(), color="red", linestyle="--",
                        label=f"Médiane={_prix_knn_s.median():.2f} £")
    _axes_ri[2].set_title("UnitPrice — KNN imputation\n(distribution préservée)", fontweight="bold")
    _axes_ri[2].set_xlabel("UnitPrice (£)")
    _axes_ri[2].legend(fontsize=8)

    plt.suptitle("🛒 Online Retail — KNN vs Médiane pour l'imputation du UnitPrice", fontsize=12, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/retail_imputation.png", dpi=120, bbox_inches="tight")
    plt.show()

    # Polars équivalent
    try:
        import polars as _pl
        _df_pl_ri = _pl.from_pandas(df_clean[["UnitPrice", "Country"]].head(100))
        _mode_pl = _df_pl_ri["Country"].drop_nulls().mode()[0]
        _df_pl_ri = _df_pl_ri.with_columns(
            _pl.col("UnitPrice").fill_null(_pl.col("UnitPrice").median()),
            _pl.col("Country").fill_null(_mode_pl),
        )
        print("\n📊 Polars — imputation en une expression :")
        print(_df_pl_ri.null_count())
    except ImportError:
        print("\n# Polars — imputation en une expression :")
        print("# df_pl.with_columns(")
        print("#     pl.col('UnitPrice').fill_null(pl.col('UnitPrice').median()),")
        print("#     pl.col('Country').fill_null(pl.col('Country').drop_nulls().mode().first()),")
        print("# )")

    df_retail_clean = _df_ri
    return (df_retail_clean,)


@app.cell
def _(df_retail_clean):
    df_retail_clean.head()
    return


@app.cell(hide_code=True)
def _retail_cat_imputation_theory(mo):
    mo.md(r"""
    ### 3.3 Imputation des variables catégorielles — Online Retail

    Les variables catégorielles ne peuvent pas être imputées par la moyenne ou la médiane (ces notions ne s'appliquent pas à des catégories textuelles). La stratégie standard est le **mode** (`most_frequent`) : on remplace chaque valeur manquante par la catégorie la plus fréquente dans la colonne.

    **Colonnes catégorielles du dataset Online Retail :**

    | Colonne | Type | Manquants | Stratégie |
    |---------|------|-----------|-----------|
    | `Country` | Pays de livraison | ~1% | `most_frequent` → **United Kingdom** (MCAR) |
    | `Description` | Libellé produit | ~5% | `most_frequent` ou suppression (MCAR) |
    | `StockCode` | Code article | ~5% | `most_frequent` ou suppression (MCAR) |
    | `InvoiceNo` | Numéro de facture | 0 | Pas de manquants |
    | `CustomerID` | Identifiant client | ~20% | Indicatrice `HasCustomerID` (MAR — ne pas imputer une ID fictive) |

    > **Pourquoi `SimpleImputer(strategy="most_frequent")` ?**
    > `sklearn` uniformise le traitement : même API pour les numériques et les catégorielles.
    > Cela facilite l'intégration dans un `Pipeline` ou un `ColumnTransformer` en production.

    > ⚠️ Pour `CustomerID`, **ne jamais imputer** une valeur fictive — une fausse ID regroupera artificiellement des clients distincts. La bonne approche est de créer une indicatrice `HasCustomerID` (section 3) et de traiter séparément les transactions anonymes.

    > ⚠️ Pour `Description`, le mode (`most_frequent`) est acceptable si le taux de manquants est < 5 %. Au-delà, préférez la suppression ou une jointure sur `StockCode` pour récupérer le libellé depuis une table produit.
    """)
    return


@app.cell
def _retail_cat_imputation(df_clean):
    from sklearn.impute import SimpleImputer as _SI_cat

    print("=== 🛒 ONLINE RETAIL — Imputation des variables catégorielles ===\n")

    _cat_features = ["StockCode", "Description", "Country"]
    _retail_cat = df_clean[_cat_features].copy()

    print("Aperçu des données catégorielles (5 premières lignes) :")
    print(_retail_cat.head().to_string())

    print("\nValeurs manquantes AVANT imputation :")
    print(_retail_cat.isnull().sum().to_string())

    # Imputation par le mode (valeur la plus fréquente)
    _imputer_cat = _SI_cat(strategy="most_frequent")
    _retail_cat_imp = _imputer_cat.fit_transform(_retail_cat)
    _retail_cat_imp = pd.DataFrame(
        _retail_cat_imp,
        columns=_retail_cat.columns,
        index=_retail_cat.index,
    )

    print("\nValeurs manquantes APRÈS imputation :")
    print(_retail_cat_imp.isnull().sum().to_string())

    print("\nValeurs imputées par colonne (mode) :")
    for _col_c, _stat_c in zip(_cat_features, _imputer_cat.statistics_):
        print(f"  {_col_c:<14} → mode = '{_stat_c}'")

    print("\nAperçu après imputation :")
    print(_retail_cat_imp.head().to_string())

    print("\n⚠️ Note : CustomerID n'est PAS imputé ici — on crée une indicatrice HasCustomerID.")
    print("   Imputer une ID fictive regrouperait artificiellement des clients distincts.")
    return


@app.cell(hide_code=True)
def _section4(mo):
    mo.md(r"""
    ---
    ## 4. Détection des outliers

    Un **outlier** est une observation significativement différente des autres, souvent c'est une valeur abérante extrême.

    Les outliers peuvent être des erreurs de saisie (à corriger), des événements rares légitimes (fraude, pic de ventes) ou des signaux importants à analyser séparément.

    ### 4.1 Méthode IQR (robuste, sans hypothèse de distribution)

    $$IQR = Q_3 - Q_1$$

    $$\text{Borne inf} = Q_1 - 1.5 \times IQR \qquad \text{Borne sup} = Q_3 + 1.5 \times IQR$$

    Toute valeur hors de cet intervalle est un outlier. Utiliser $3 \times IQR$ pour une détection plus conservatrice.

    ### 4.2 Z-score (pour distributions approximativement normales)

    $$z_i = \frac{x_i - \bar{x}}{\sigma}$$

    Avec $\bar{x}$ : la moyenne
    Et $\sigma$ : L'écar type

    **Règle :** Si $|z_i| > 3$, la valeur est un outlier (règle des 3 sigma). Cette méthode suppose une distribution normale — elle est sensible aux outliers eux-mêmes (problème de circularité).

    ### 4.3 Isolation Forest

    Algorithme non supervisé basé sur des arbres de décision aléatoires.
    Les outliers sont **isolés plus rapidement** que les points normaux car ils sont dans des régions peu denses.
    Particulièrement adapté aux données **haute dimension** ou aux distributions non normales.
    """)
    return


@app.cell
def _demo_outliers(df_clean):
    from scipy import stats as _scipy_stats
    from sklearn.ensemble import IsolationForest

    print("=== DÉTECTION DES OUTLIERS — Quantity ===\n")
    _qty5 = df_clean["Quantity"].dropna()

    # IQR
    _Q1_5 = _qty5.quantile(0.25)
    _Q3_5 = _qty5.quantile(0.75)
    _IQR5 = _Q3_5 - _Q1_5
    _b_inf5 = _Q1_5 - 1.5 * _IQR5
    _b_sup5 = _Q3_5 + 1.5 * _IQR5
    _out_iqr5 = (_qty5 < _b_inf5) | (_qty5 > _b_sup5)
    print(f"IQR = {_IQR5:.1f}  |  Q1 = {_Q1_5:.1f}  |  Q3 = {_Q3_5:.1f}")
    print(f"Bornes IQR : [{_b_inf5:.1f}, {_b_sup5:.1f}]")
    print(f"Outliers IQR : {_out_iqr5.sum()} ({_out_iqr5.sum()/len(_qty5)*100:.1f}%)")

    # Z-score
    _z5 = np.abs(_scipy_stats.zscore(_qty5))
    _out_z5 = _z5 > 3
    print(f"\nOutliers Z-score (|z|>3) : {_out_z5.sum()} ({_out_z5.sum()/len(_qty5)*100:.1f}%)")

    # Isolation Forest
    _iso5 = IsolationForest(contamination=0.05, random_state=42)
    _pred5 = _iso5.fit_predict(_qty5.values.reshape(-1, 1))
    _out_iso5 = _pred5 == -1
    print(f"Outliers Isolation Forest : {_out_iso5.sum()} ({_out_iso5.sum()/len(_qty5)*100:.1f}%)")

    qty_clean = _qty5
    out_iqr = _out_iqr5
    b_inf = _b_inf5
    b_sup = _b_sup5
    out_iso = _out_iso5
    Q1_qty = _Q1_5
    Q3_qty = _Q3_5
    return b_inf, b_sup, out_iqr, out_iso, qty_clean


@app.cell
def _plot_outliers(b_inf, b_sup, out_iqr, out_iso, qty_clean):
    _fig6, _axes6 = plt.subplots(1, 3, figsize=(16, 5))

    # Boxplot
    _ax6a = _axes6[0]
    _ax6a.boxplot(qty_clean, vert=True, patch_artist=True,
                  boxprops=dict(facecolor="#3498db", alpha=0.6),
                  flierprops=dict(marker="o", markerfacecolor="#e74c3c", markeredgecolor="#e74c3c", markersize=4, alpha=0.7))
    _ax6a.set_title("Boxplot", fontweight="bold")
    _ax6a.set_ylabel("Quantity")
    _ax6a.axhline(b_inf, color="orange", linestyle="--", alpha=0.7, label="Bornes IQR")
    _ax6a.axhline(b_sup, color="orange", linestyle="--", alpha=0.7)
    _outlier_patch = plt.matplotlib.lines.Line2D(
        [], [], marker="o", color="w", markerfacecolor="#e74c3c",
        markersize=6, label=f"Outliers IQR (n={out_iqr.sum()})")
    _iqr_line = plt.matplotlib.lines.Line2D(
        [], [], color="orange", linestyle="--", label="Bornes IQR")
    _ax6a.legend(handles=[_outlier_patch, _iqr_line], fontsize=8)

    # Distribution + seuils IQR
    _ax6b = _axes6[1]
    _in6 = qty_clean[~out_iqr]
    _out6 = qty_clean[out_iqr]
    _ax6b.hist(_in6, bins=40, color="#3498db", alpha=0.7, label="Normal")
    _ax6b.hist(_out6, bins=20, color="#e74c3c", alpha=0.9, label=f"Outliers IQR (n={out_iqr.sum()})")
    _ax6b.axvline(b_inf, color="orange", linestyle="--", linewidth=2, label="Bornes IQR")
    _ax6b.axvline(b_sup, color="orange", linestyle="--", linewidth=2)
    _ax6b.set_title("Distribution des quantités", fontweight="bold")
    _ax6b.set_xlabel("Quantity")
    _ax6b.set_xlim(-100, 250)
    _ax6b.legend(fontsize=8)

    # Comparaison méthodes
    _ax6c = _axes6[2]
    _methods6 = ["IQR", "Z-score\n(|z|>3)", "Isolation\nForest"]
    _counts6 = [out_iqr.sum(),
                int((np.abs((qty_clean - qty_clean.mean()) / qty_clean.std()) > 3).sum()),
                out_iso.sum()]
    _bars6 = _ax6c.bar(_methods6, _counts6, color=["#3498db", "#2ecc71", "#9b59b6"], alpha=0.8, edgecolor="white")
    _ax6c.set_title("Outliers détectés par méthode", fontweight="bold")
    _ax6c.set_ylabel("Nombre")
    for _b6, _c6 in zip(_bars6, _counts6):
        _ax6c.text(_b6.get_x() + _b6.get_width()/2, _b6.get_height() + 0.5,
                   str(_c6), ha="center", fontweight="bold")

    plt.suptitle("Comparaison des méthodes de détection d'outliers — Quantity", fontsize=12, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/outliers.png", dpi=120, bbox_inches="tight")
    plt.show()
    return


@app.cell(hide_code=True)
def _section5(mo):
    mo.md(r"""
    ---
    ## 5. Normalisation et Standardisation

    Les algorithmes ML sont sensibles aux **différences d'échelle**.
    Une variable `salaire` (0–100 000 €) dominerait une variable `âge` (0–100 ans) sans mise à l'échelle.

    ### 5.1 Normalisation Min-Max (MinMaxScaler)

    Ramène toutes les valeurs dans $[0, 1]$ :

    $$x_{\text{norm}} = \frac{x - x_{\min}}{x_{\max} - x_{\min}}$$

    **Avantages :** Préserve la distribution. Utile pour les réseaux de neurones (valeurs bornées).
    **Inconvénients :** Très sensible aux outliers — un outlier extrême comprime toutes les autres valeurs.

    ### 5.2 Standardisation Z-score (StandardScaler)

     Transforme les données à une distribution centrée et réduite : $\mu = 0$, $\sigma = 1$. Chaque valeur devient le nombre d'écarts-types qui la séparent de la moyenne.

    $$z_i = \frac{x_i - \bar{x}}{\sigma}$$


    **Avantages :** Standard pour SVM, régression logistique, PCA.
    **Inconvénients :** Ne garantit pas une plage fixe.

    ### 5.3 RobustScaler

    Utilise la **médiane et l'IQR** — robuste aux outliers :

    $$x_{\text{robust}} = \frac{x - \text{médiane}(x)}{IQR}$$

    > ⚠️ **Règle critique :** Toujours `.fit()` sur le **train set uniquement**, puis `.transform()` sur le test set. Fitter sur le test set = *data leakage*.
    """)
    return


@app.cell
def _demo_scaling(df_clean):
    from sklearn.preprocessing import MinMaxScaler, StandardScaler, RobustScaler

    _prix7 = df_clean["UnitPrice"].dropna()
    _prix7 = _prix7[_prix7 < _prix7.quantile(0.99)].values.reshape(-1, 1)

    _scalers7 = {
        "Original": _prix7.flatten(),
        "MinMaxScaler\n[0, 1]": MinMaxScaler().fit_transform(_prix7).flatten(),
        "StandardScaler\n(μ=0, σ=1)": StandardScaler().fit_transform(_prix7).flatten(),
        "RobustScaler\n(médiane/IQR)": RobustScaler().fit_transform(_prix7).flatten(),
    }

    _fig7, _axes7 = plt.subplots(1, 4, figsize=(16, 4))
    _cols7 = ["#95a5a6", "#3498db", "#2ecc71", "#e67e22"]

    for _ax7, (_lbl7, _d7), _c7 in zip(_axes7, _scalers7.items(), _cols7):
        _ax7.hist(_d7, bins=30, color=_c7, alpha=0.8, edgecolor="white")
        _ax7.axvline(np.mean(_d7), color="red", linestyle="--", linewidth=1.5,
                     label=f"μ={np.mean(_d7):.2f}")
        _ax7.axvline(np.median(_d7), color="purple", linestyle=":", linewidth=1.5,
                     label=f"med={np.median(_d7):.2f}")
        _ax7.set_title(_lbl7, fontweight="bold", fontsize=10)
        _ax7.legend(fontsize=7)
        _info7 = f"min={_d7.min():.2f}\nmax={_d7.max():.2f}\nσ={_d7.std():.2f}"
        _ax7.text(0.97, 0.97, _info7, transform=_ax7.transAxes, va="top", ha="right",
                  fontsize=8, bbox=dict(boxstyle="round", facecolor="white", alpha=0.8))

    plt.suptitle("Comparaison des méthodes de mise à l'échelle — UnitPrice", fontsize=12, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/scaling.png", dpi=120, bbox_inches="tight")
    plt.show()
    print("✅ La distribution est préservée — seule l'échelle change")
    return


@app.cell
def _(df_clean):
    # Orignal distribution 

    _df_rs = df_clean[["Quantity"]].copy()
    plt.hist(_df_rs["Quantity"], bins=100, color="#95a5a6", alpha=0.8, edgecolor="white")
    return


@app.cell
def _retail_scaling(df_clean):
    from sklearn.preprocessing import MinMaxScaler as _MMS_r, StandardScaler as _SS_r, RobustScaler as _RS_r

    print("=== 🛒 ONLINE RETAIL — Mise à l'échelle de Quantity et UnitPrice ===\n")
    _df_rs = df_clean[["Quantity", "UnitPrice"]].copy()
    _df_rs = _df_rs[(_df_rs["Quantity"] > 0) & (_df_rs["UnitPrice"] > 0)]

    print(f"Quantity  : min={_df_rs['Quantity'].min():.0f}   max={_df_rs['Quantity'].max():.0f}   "
          f"médiane={_df_rs['Quantity'].median():.1f}  σ={_df_rs['Quantity'].std():.1f}")
    print(f"UnitPrice : min={_df_rs['UnitPrice'].min():.2f}  max={_df_rs['UnitPrice'].max():.2f}  "
          f"médiane={_df_rs['UnitPrice'].median():.2f}  σ={_df_rs['UnitPrice'].std():.2f}")
    print("\n→ Les deux variables ont des outliers (grandes quantités B2B, articles premium)")
    print("→ RobustScaler recommandé — moins sensible aux grosses commandes et aux prix extrêmes")

    _scalers_r = {
        "MinMaxScaler": _MMS_r(),
        "StandardScaler": _SS_r(),
        "RobustScaler": _RS_r(),
    }
    _fig_rs, _axes_rs = plt.subplots(2, 3, figsize=(15, 7))
    _cols_rs = ["#3498db", "#2ecc71", "#e67e22"]

    for _j, ((_lbl_rs, _scaler_rs), _col_rs) in enumerate(zip(_scalers_r.items(), _cols_rs)):
        _scaled_r = _scaler_rs.fit_transform(_df_rs)
        for _i, _feat_r in enumerate(["Quantity", "UnitPrice"]):
            _ax_rs = _axes_rs[_i, _j]
            _ax_rs.hist(_scaled_r[:, _i], bins=100, color=_col_rs, alpha=0.8, edgecolor="white")
            _ax_rs.axvline(np.mean(_scaled_r[:, _i]), color="red", linestyle="--", linewidth=1.5,
                           label=f"μ={np.mean(_scaled_r[:, _i]):.2f}")
            _ax_rs.set_title(f"{_lbl_rs}\n{_feat_r}", fontweight="bold", fontsize=9)
            _info_rs = f"min={_scaled_r[:, _i].min():.2f}\nmax={_scaled_r[:, _i].max():.2f}"
            _ax_rs.text(0.97, 0.97, _info_rs, transform=_ax_rs.transAxes,
                        va="top", ha="right", fontsize=8,
                        bbox=dict(boxstyle="round", facecolor="white", alpha=0.8))
            _ax_rs.legend(fontsize=7)

    plt.suptitle("🛒 Online Retail — Scaling de Quantity et UnitPrice (outliers B2B et articles premium)\n"
                 "→ RobustScaler moins sensible aux grosses commandes grossistes",
                 fontsize=11, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/retail_scaling.png", dpi=120, bbox_inches="tight")
    plt.show()

    # Polars équivalent
    try:
        import polars as _pl
        _df_pl_rs = _pl.from_pandas(_df_rs)
        _median_rs = _df_pl_rs["UnitPrice"].median()
        _iqr_rs = _df_pl_rs["UnitPrice"].quantile(0.75) - _df_pl_rs["UnitPrice"].quantile(0.25)
        _df_pl_rs = _df_pl_rs.with_columns(
            ((_pl.col("UnitPrice") - _median_rs) / _iqr_rs).alias("UnitPrice_robust")
        )
        print("\n📊 Polars — RobustScaler manuel sur UnitPrice :")
        print(_df_pl_rs.select(["UnitPrice", "UnitPrice_robust"]).describe())
    except ImportError:
        print("\n# Polars — RobustScaler manuel :")
        print("# median = df_pl['UnitPrice'].median()")
        print("# iqr = df_pl['UnitPrice'].quantile(0.75) - df_pl['UnitPrice'].quantile(0.25)")
        print("# df_pl.with_columns(((pl.col('UnitPrice') - median) / iqr).alias('UnitPrice_robust'))")
    return


@app.cell(hide_code=True)
def _section_td(mo):
    mo.md(r"""
    ---
    ## TD à rendre — Exercices de nettoyage sur le dataset Titanic

    **A Rendre au plus tard Mercredi 01/06/2026 avant 23h 59.**

    **Mail** : mboup.djibril@ugb.edu.sn

    ### Contexte métier

    Vous êtes data analyst pour une compagnie maritime. On vous fournit le manifeste des passagers du RMS Titanic (1 309 passagers, 14 colonnes). Votre mission : **nettoyer et préparer les données dans le but d'analyser les chances de survie de ce nauffrage**.

    | Colonne | Description |
    |---------|-------------|
    | `pclass` | Classe du billet (1 = 1ère, 2 = 2ème, 3 = 3ème) |
    | `survived` | Survie (1 = oui, 0 = non) |
    | `name` | Nom complet du passager |
    | `sex` | Sexe |
    | `age` | Âge en années |
    | `sibsp` | Nombre de frères/sœurs ou conjoint(e)s à bord |
    | `parch` | Nombre de parents/enfants à bord |
    | `ticket` | Numéro de ticket |
    | `fare` | Tarif payé (£) |
    | `cabin` | Numéro de cabine |
    | `embarked` | Port d'embarquement (S = Southampton, C = Cherbourg, Q = Queenstown) |
    | `boat` | Numéro du canot de sauvetage (si survivant) |
    | `body` | Numéro d'identification du corps (si décédé retrouvé) |
    | `home.dest` | Ville/destination de résidence |

    ---

    ### Exercice 1 — Audit du dataset

    Chargez le dataset et analysez son état de santé :

    ```python
    df_titanic = pd.read_csv("datasets/raw/titanic.csv")
    ```

    Répondez aux questions suivantes :

    1. Combien de lignes et de colonnes contient le dataset ?
    2. Quelle colonne présente le taux de valeurs manquantes le plus élevé ? Pourquoi est-ce logique ?
    3. Quels sont les types de données de chaque colonne ? Y a-t-il des incohérences ?
    4. Combien de doublons y a-t-il ?

    ---

    ### Exercice 2 — Mécanismes de données manquantes

    Pour chaque colonne ci-dessous, identifiez le mécanisme (MCAR / MAR / MNAR) et proposez une stratégie d'imputation :

    | Colonne | % Manquants | Contexte | Mécanisme ? | Votre stratégie ? |
    |---------|------------|----------|-------------|-------------------|
    | `age` | ~20% | Âge non renseigné sur le manifeste | ? | ? |
    | `cabin` | ~77% | Numéro de cabine | ? | ? |
    | `embarked` | <1% | Port d'embarquement | ? | ? |
    | `fare` | <1% | Tarif du billet | ? | ? |
    | `boat` | ~63% | Numéro du canot de sauvetage | ? | ? |
    | `body` | ~90% | Numéro du corps retrouvé | ? | ? |

    > **Indice pour `boat` et `body` :** réfléchissez à ce que signifie une valeur manquante dans ces colonnes par rapport à la variable `survived`.

    ---

    ### Exercice 3 — Comparer deux stratégies de nettoyage

    **Stratégie A :** Supprimer toutes les lignes avec au moins une valeur manquante (sur les colonnes `age`, `embarked`, `fare`).

    **Stratégie B :**
    - Imputer `age` par la **médiane par classe et par sexe** (MAR)
    - Imputer `embarked` par le **mode** (MCAR, 2 valeurs seulement)
    - Imputer `fare` par la **médiane** (MCAR, 1 valeur seulement)
    - Créer une indicatrice `has_cabin` (1 si `cabin` renseigné, 0 sinon)

    Pour chaque stratégie, calculez :
    - Le nombre de passagers conservés et le % de données perdues
    - Le taux de survie global (`survived.mean()`)
    - La répartition homme/femme conservée

    > Comparez les deux taux de survie : la stratégie A introduit-elle un biais ?

    ---

    ### Exercice 4 — Détection et traitement des outliers

    1. Appliquez la méthode IQR sur la colonne `fare`. Combien d'outliers détectez-vous ?
    2. Ces outliers sont-ils des erreurs ou des valeurs légitimes ? Justifiez.
    3. Comparez avec le Z-score (seuil |z| > 3). Y a-t-il une différence ?
    4. Quelle stratégie adopteriez-vous : suppression, remplacement, ou conservation ?

    ---

    ### Exercice 5 — Discussion

    1. Les données manquantes dans `body` sont-elles MCAR, MAR ou MNAR ? Expliquez le lien avec `survived`.
    2. Si on supprime tous les passagers sans `cabin`, quel biais introduit-on par rapport à `pclass` ?
    3. Faut-il conserver les colonnes `boat`, `body`, `ticket`, `name` pour un modèle de prédiction ? Pourquoi ?
    4. L'imputation de `age` par la médiane globale vs la médiane par `pclass`+`sex` — quelle différence observez-vous ?
    """)
    return




@app.cell(hide_code=True)
def _section7(mo):
    mo.md(r"""
    ---
    ## 7. Profiling automatique — ydata-profiling

    `ydata-profiling` génère un rapport HTML complet du dataset en une seule ligne de code.

    ```python
    from ydata_profiling import ProfileReport

    # Rapport complet
    profile = ProfileReport(df, title="Online Retail — Rapport de profiling", explorative=True)
    profile.to_notebook_iframe()  # afficher dans le notebook
    profile.to_file("rapport.html")  # exporter en HTML

    # Mode minimal (plus rapide sur grands datasets)
    profile = ProfileReport(df, minimal=True)

    # Sur échantillon pour datasets > 100 000 lignes
    profile = ProfileReport(df.sample(10_000), title="Profiling sur échantillon")
    ```

    ### Contenu du rapport

    | Section | Informations |
    |---------|-------------|
    | Vue d'ensemble | Dimensions, types, % manquants, doublons |
    | Par variable | Distribution, statistiques, alertes (skewness, corrélations) |
    | Corrélations | Pearson, Spearman, Cramér's V (catégorielles) |
    | Interactions | Scatter plots entre paires de variables |
    | Échantillon | Premières et dernières lignes |
    """)
    return


@app.cell
def _profiling_simplifie(df_clean):
    print("=== PROFILING SIMPLIFIÉ ===\n")
    _df10 = df_clean.copy()
    _n10, _c10 = _df10.shape
    print(f"{'─'*55}")
    print(f"  📐 Dimensions  : {_n10:,} lignes × {_c10} colonnes")
    print(f"  ⚠️  Manquants   : {_df10.isnull().sum().sum():,} ({_df10.isnull().sum().sum()/(_n10*_c10)*100:.1f}%)")
    print(f"  🔁 Doublons    : {_df10.duplicated().sum()}")
    print(f"{'─'*55}\n")
    print(f"{'Colonne':<15} {'Type':<12} {'Manquants':<14} {'Uniques':<10} {'Alertes'}")
    print("─" * 65)
    for _col10 in _df10.columns:
        _dtype10 = str(_df10[_col10].dtype)
        _manq10 = _df10[_col10].isnull().sum()
        _pct10 = _manq10 / len(_df10) * 100
        _uniq10 = _df10[_col10].nunique()
        _alerts10 = []
        if _pct10 > 30: _alerts10.append("⚠️ CRITIQUE")
        elif _pct10 > 10: _alerts10.append("⚡ Modéré")
        if _uniq10 == 1: _alerts10.append("🚫 Constante")
        _m10str = f"{_manq10} ({_pct10:.0f}%)" if _manq10 > 0 else "0"
        print(f"{_col10:<15} {_dtype10:<12} {_m10str:<14} {_uniq10:<10} {', '.join(_alerts10)}")
    return


@app.cell
def _():
    return


@app.cell
def _(df_clean):
    try:
        from ydata_profiling import ProfileReport
        _df_profiling = df_clean.copy()
        profile = ProfileReport(_df_profiling, title="Online Retail — Rapport de profiling", explorative=True)
        profile.to_notebook_iframe()
        profile.to_file("rapport_profiling.html")
        print("✅ Rapport exporté → rapport_profiling.html")
    except ImportError:
        print("⚠️ ydata-profiling non installé — lancez : uv add ydata-profiling setuptools")
    except Exception as e:
        print(f"⚠️ Erreur profiling : {e}")
    return


@app.cell(hide_code=True)
def _section_column_transformer(mo):
    mo.md(r"""
    ---
    ## 9. Pipeline complet — ColumnTransformer

    En production, on sépare le traitement des variables **numériques** et **catégorielles** dans des sous-pipelines distincts, assemblés par `ColumnTransformer`.

    ```
    ColumnTransformer
    ├── num_pipeline     → [SimpleImputer(median)] → [RobustScaler()]
    └── cat_pipeline     → [SimpleImputer(mode)]   → [OneHotEncoder(drop='first')]
    ```

    Cette architecture garantit :
    - **Reproductibilité** — un seul objet `full_pipeline` à sauvegarder avec `joblib`
    - **Pas de data leakage** — `fit()` sur train, `transform()` sur test
    - **Extensibilité** — ajouter une étape = insérer un transformateur dans le pipeline

    ```python
    from sklearn.pipeline import Pipeline
    from sklearn.compose import ColumnTransformer
    from sklearn.impute import SimpleImputer
    from sklearn.preprocessing import RobustScaler, OneHotEncoder

    num_pipeline = Pipeline([
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler",  RobustScaler()),
    ])

    cat_pipeline = Pipeline([
    ("imputer", SimpleImputer(strategy="most_frequent"))
    ])

    full_pipeline = ColumnTransformer([
    ("num", num_pipeline, num_cols),
    ("cat", cat_pipeline, cat_cols),
    ])

    # Toujours : fit sur TRAIN uniquement
    X_prepared = full_pipeline.fit_transform(X)
    #X_test_prepared  = full_pipeline.transform(X_test)   #ML Use case ← pas de fit_transform pour les données pour éviter le data leakage!
    ```
    """)
    return


@app.cell
def _demo_column_transformer(df_clean):
    from sklearn.pipeline import Pipeline as _Pipe
    from sklearn.compose import ColumnTransformer
    from sklearn.impute import SimpleImputer as _SI_ct
    from sklearn.preprocessing import RobustScaler as _RS_ct, OneHotEncoder as _OHE_ct
    from sklearn.model_selection import train_test_split
    import os

    _df_ct = df_clean[["Quantity", "UnitPrice", "CustomerID", "Country"]].copy()
    _df_ct = _df_ct[_df_ct["Quantity"] > 0].reset_index(drop=True)
    _df_ct.rename(columns={"Quantity": "Quantity_capped", "UnitPrice": "UnitPrice_capped"}, inplace=True)
    _df_ct["HasCustomerID"] = _df_ct["CustomerID"].notna().astype(int)
    _df_ct["Country"] = _df_ct["Country"].fillna("United Kingdom")


    _num_cols_ct = ["Quantity_capped", "UnitPrice_capped", "HasCustomerID"]
    _cat_cols_ct = ["Country"]

    _X = _df_ct[_num_cols_ct + _cat_cols_ct]


    _num_pipe = _Pipe([
        ("imputer", _SI_ct(strategy="median")),
        ("scaler",  _RS_ct()),
    ])
    _cat_pipe = _Pipe([
        ("imputer", _SI_ct(strategy="most_frequent")),
    ])
    _full_pipeline = ColumnTransformer([
        ("num", _num_pipe, _num_cols_ct),
        ("cat", _cat_pipe, _cat_cols_ct),
    ])

    _X_prep = _full_pipeline.fit_transform(_X)

    print("=== COLUMN TRANSFORMER — Pipeline complet ===\n")
    print(f"Prep Data : {_X.shape} → {_X_prep.shape}")

    print(f"\nValeurs manquantes après pipeline : {(_X_prep != _X_prep).sum()}")

    #_ohe_cats = _full_pipeline.named_transformers_["cat"]["encoder"].categories_[0][1:]
    _feat_names = _num_cols_ct + _cat_cols_ct
    _df_prep_preview = pd.DataFrame(_X_prep, columns=_feat_names).round(3)
    print(f"\n📊 Aperçu du dataset préparé (5 premières lignes) :")
    print(_df_prep_preview.head())
    print("\n Pipeline prêt — fitter une seule fois, sauvegarder en parquet")
    os.makedirs("datasets/prepared/parquet",  exist_ok=True)
    _df_prep_preview.to_parquet("datasets/prepared/parquet/retail_online.parquet")
    print("\n✅ Data save in Parquet")
    return


@app.cell(hide_code=True)
def _section_sampling(mo):
    mo.md(r"""
    ---
    ## 10. Échantillonnage

    ### Population
    C'est un ensemble de tous les individus, objets ou unités destinées à une analyse ou l'objet d'une étude statistiques.

    - Les habitants d'un pays d'une régions, les entrepises agricoles, les clients d'une sociéte, les passagers d'un crashe etc.

    ### Un échantillon
    Un échantillon est un groupe relativement petit et choisi scientifiquement de manière à représenter le plus fidèlement possible une population.

    Au lieu d’examiner l’ensemble de la population, on étudie une partie ou un sous-ensemble de cette population qui est représentatif et à partir duquel on peut tirer des conclusions pour l’ensemble de cette population.

    La statistique inférentielle permet, à l’aide des probabilités, de généraliser les conclusions issues d’un échantillon pour l’ensemble de la population avec un certain degré de certitude.


    - Collecter toutes les données d'une population n'est pas possible, d'où la necessité de travaillé dans un échantillon
    - Traiter et analyser sur toutes les données d'une population peut être une tâche fastidieuse.
    - Faire une étude sur un datasets contenant des millions de lignes est lent et coûteux
    - Certaines analyses (EDA, tests d'imputation) n'ont pas besoin de toutes les données


    L'**échantillonnage** est donc une étape importante sur l'analyse statistique des données. Un mauvais échantillon ne produit pas seulement de mauvaises statistiques —
    il **biaise** les décisions prise à partir des données.

    ⚛️ Nous avons deux types d'échantillonage : les **échantillons probabilistes** et les **échantillons non-probabilistes**.

    **Echantillon probabiliste** :
    Les individues ou objets sont choisis selon une procédure où la sélection est aléatoire.

    **Exemples** :
    - Échantillonnage aléatoire simple
    - Échantillonnage aléatoire stratifié
    - Échantillonnage par grappes
    - Échantillonnage systématique

    **Echantillons non-probabiliste** :
    Les échantillons sont sélectionnés en fonction du jugement ou de la facilité d'accès aux données. Cela dépend largement des compétences du chercheur en matière de sélection d'échantillons.

    **Exemples** :
    - Échantillonnage de convenance
    - Échantillonnage ciblé (ou échantillonnage dirigé / par choix raisonné)
    - Échantillonnage de volontaires (ou échantillonnage à réponse volontaire)
    - Échantillonnage en boule de neige (ou par réseau)

    ### Position dans le pipeline

    ```
    Collecte → [Échantillonnage] → Audit qualité → Nettoyage → Préparation → Modélisation
                     ↑
              Cette section couvre cette étape
    ```

    ## 10.1 — Le dataset comme population de référence

    Le dataset Online Retail (UCI ML Repository) joue ici le rôle de **population** :
    c'est l'ensemble des $N$ transactions disponibles, avec tous leurs défauts
    (doublons, manquants, outliers) déjà observés aux sections précédentes.

    ### Vocabulaire

    | Terme | Définition | Dans notre cas |
    |-------|-----------|----------------|
    | **Population** | Ensemble complet des individus d'intérêt (taille $N$) | Toutes les transactions du dataset brut |
    | **Échantillon** | Sous-ensemble tiré de la population (taille $n$) | Les lignes sélectionnées pour tester le pipeline |
    | **Taux de sondage** | $f = n/N$ | Typiquement $f < 10\%$ en pratique |
    | **Paramètre** | Vraie valeur dans la population ($\mu$, $\sigma$…) | UnitPrice moyen, taux de manquants réel |
    | **Estimateur** | Valeur calculée sur l'échantillon ($\bar{x}$, $\hat{p}$…) | Ce que le pipeline "voit" et sur quoi il se calibre |

    La cellule suivante affiche les paramètres réels de la population — c'est ce que
    les différentes stratégies d'échantillonnage vont tenter de reproduire fidèlement.
    """)
    return


@app.cell(hide_code=True)
def _srs_theory(mo):
    mo.md(r"""
    ## 10.2 Échantillonnage aléatoire simple — tester le pipeline sur un sous-ensemble

    Avant de dérouler le pipeline complet sur les centaines de milliers de lignes d'une
    population, on le développe et le valide sur un **échantillon représentatif**.
    L'approche la plus naturelle : l'**échantillonnage aléatoire simple** (SRS).

    *« In a simple random sample (SRS), each sampling unit and every possible combination of sampling units has an equal chance of selection »* (Sheskin, 1985)

    Chaque transaction a la même probabilité $\dfrac{1}{N}$ d'être sélectionnée.

    ### **Sans remise**

    Chaque ligne ne peut apparaître qu'une seule fois dans l'échantillon.
    C'est le bon choix pour développer et valider les étapes de nettoyage :
    on s'assure que l'échantillon reproduit fidèlement la distribution des défauts
    (taux de manquants, proportion d'outliers, diversité des pays).

    En pratique : `df.sample(n=100, replace=False)`.

    ### **Avec remise**  — utile pour la méthode bootstrap (§ 12.7)

    Une même ligne peut apparaître plusieurs fois. Moins utilisé pour la préparation
    directe des données, mais **fondamental pour quantifier l'incertitude** sur les
    statistiques calculées après nettoyage.

    ### Propriétés de l'estimateur

    Sur un SRS, la moyenne empirique $\bar{x}$ est **sans biais** ($\mathbb{E}[\bar{x}] = \mu$)
    avec une variance $\text{Var}(\bar{x}) = \sigma^2/n$.
    """)
    return


@app.cell
def _():
    import openpyxl
    path_retail = "datasets/online_retail/Online_Retail.xlsx"
    df_ret_orig = pd.read_excel(path_retail, 
                            parse_dates=['InvoiceDate'],
                            dtype={'CustomerID': str},   # éviter la conversion float → perte du zéro de tête
                            na_values=['', 'NA', '?'],
                           )
    print(df_ret_orig.info())
    return (df_ret_orig,)


@app.cell
def _retail_pour_sampling(df_ret_orig):
    """Paramètres réels de la population — référence pour évaluer chaque stratégie d'échantillonnage."""
    retail = df_ret_orig.dropna(subset=["Quantity", "UnitPrice", "Country"]).copy()
    retail = retail[retail["UnitPrice"] > 0].reset_index(drop=True)
    retail["Revenue"] = retail["Quantity"] * retail["UnitPrice"]

    print("=== POPULATION — Online Retail (données brutes, avant pipeline complet) ===")
    print()
    print(f"  Taille totale         : N = {len(retail):,} transactions")
    print(f"  UnitPrice moyen       : μ = {retail['UnitPrice'].mean():.2f} £")
    print(f"  UnitPrice médian      : {retail['UnitPrice'].median():.2f} £  "
          f"(très différent de la moyenne → distribution asymétrique)")
    print(f"  Taux de manquants     : CustomerID = {retail['CustomerID'].isna().mean():.1%}")
    print(f"  Transactions UK       : {(retail['Country']=='United Kingdom').mean():.1%}  "
          f"← déséquilibre géographique fort")
    print(f"  Pays distincts        : {retail['Country'].nunique()}")
    print()
    print("  Ces paramètres sont les 'vraies' valeurs que chaque stratégie d'échantillonnage")
    print("  doit reproduire fidèlement — sans quoi le pipeline de nettoyage sera mal calibré.")
    return (retail,)


@app.cell
def _srs_demo(retail):
    """Vérifier que le SRS reproduit fidèlement les défauts de qualité de la population."""
    _n = 100 #  Taille de notre échantillon
    _srs = retail.sample(n=_n, replace=False, random_state=42) # `random_state` sert d'initialisation au générateur de nombres aléatoires afin de garantir la parfaite reproductibilité des résultats. Si vous exécutez le code plusieurs fois avec la même initialisation entière sur le même jeu de données, vous obtiendrez toujours exactement les mêmes lignes.

    def _ecart(estime, vrai):
        return f"{estime:.2f}  (écart : {(estime-vrai)/vrai*100:+.1f}%)"

    print("=== SRS (n=100) — L'échantillon reproduit-il les défauts de la population ? ===")
    print()
    print("  Statistiques de prix :")
    print(f"    UnitPrice moyen  — pop. {retail['UnitPrice'].mean():.2f} £  |  SRS : {_ecart(_srs['UnitPrice'].mean(), retail['UnitPrice'].mean())}")
    print(f"    UnitPrice médian — pop. {retail['UnitPrice'].median():.2f} £  |  SRS : {_ecart(_srs['UnitPrice'].median(), retail['UnitPrice'].median())}")
    print()
    print("  Défauts de qualité (ce que le pipeline devra corriger) :")
    print(f"    Taux manquants CustomerID — pop. {retail['CustomerID'].isna().mean():.1%}  |  SRS : {_srs['CustomerID'].isna().mean():.1%}")
    print(f"    Quantités négatives       — pop. {(retail['Quantity']<0).mean():.1%}  |  SRS : {(_srs['Quantity']<0).mean():.1%}")
    print(f"    Transactions UK           — pop. {(retail['Country']=='United Kingdom').mean():.1%}  |  SRS : {(_srs['Country']=='United Kingdom').mean():.1%}")
    print()
    print("  ► Le SRS reproduit bien les proportions de défauts.")
    print("    Un pipeline calibré sur ce SRS (médiane pour l'imputation, IQR pour les")
    print("    outliers) produira des règles transposables à la population entière.")
    print()
    print("  ► Mais : le SRS ne garantit pas la présence de chaque pays.")
    print("    Si certains marchés sont absents de l'échantillon, leurs défauts spécifiques")
    print("    (ex. CustomerID manquant à 40% en Allemagne vs 18% au UK) seront ignorés.")
    return


@app.cell
def _(mo):
    _title = mo.md("## 10.3 Échantillonnage stratifié")

    _principe = mo.md(r"""
    ### Principe

    On divise la population en **strates** homogènes selon une variable connue
    (sexe, classe d'âge, catégorie socio-professionnelle, classe du billet…),
    puis on tire un échantillon aléatoire simple (SRS) **indépendamment dans chaque strate**.

    ### Allocation proportionnelle

    Soit $H$ strates de tailles $N_1, N_2, \ldots, N_H$ avec $\sum_{h=1}^{H} N_h = N$.
    Pour un échantillon total de taille $n$, chaque strate $h$ reçoit :

    $$n_h = n \times \frac{N_h}{N}$$

    Les proportions de la population sont ainsi **exactement reproduites** dans l'échantillon.
    """)
    _output = mo.vstack([_title, _principe])
    _output
    return


@app.cell
def _stratified_theory(mo):
    _title = mo.md("## 10.3 Échantillonnage stratifié")

    _principe = mo.md(r"""
    ### Principe

    On divise la population en **strates** homogènes selon une variable connue
    (sexe, classe d'âge, catégorie socio-professionnelle, classe du billet…),
    puis on tire un échantillon aléatoire simple (SRS) **indépendamment dans chaque strate**.

    ### Allocation proportionnelle

    Soit $H$ strates de tailles $N_1, N_2, \ldots, N_H$ avec $\sum_{h=1}^{H} N_h = N$.
    Pour un échantillon total de taille $n$, chaque strate $h$ reçoit :

    $$n_h = n \times \frac{N_h}{N}$$

    Les proportions de la population sont ainsi **exactement reproduites** dans l'échantillon.
    """)

    _example = mo.callout(
        mo.md(r"""
    **Exemple — population de 5 000 habitants (hommes / femmes)**

    | Groupe | Effectif $N_h$ | Part $N_h / N$ | Taille échantillon $n_h$ (pour $n = 100$) |
    |--------|---------------|---------------|------------------------------------------|
    | Hommes | 2 750 | 55 % | $100 \times 0{,}55 = \mathbf{55}$ |
    | Femmes | 2 250 | 45 % | $100 \times 0{,}45 = \mathbf{45}$ |
    | **Total** | **5 000** | **100 %** | **100** |

    Les mêmes proportions (55 % / 45 %) sont garanties dans l'échantillon.
    """),
        kind="info",
    )

    _avantages = mo.md(r"""
    ### Avantages et limites

    | | Détail |
    |---|---|
    | ✅ **Plus précis que le SRS** | Quand la variable de stratification est corrélée à la variable étudiée |
    | ✅ **Représentation garantie** | Chaque sous-population est présente, même les groupes rares |
    | ✅ **Variance réduite** | $\text{Var}(\bar{x}_{\text{strat}}) \leq \text{Var}(\bar{x}_{\text{SRS}})$ |
    | ❌ **Nécessite une variable de stratification connue** | Inapplicable si on ne connaît pas la structure de la population |
    | ❌ **Strates trop petites** | Si $N_h$ est très faible, $n_h$ peut être < 1 → nécessite une allocation minimale |
    """)

    _retail_context = mo.md(r"""
    ### 🛒 Application Online Retail — stratifier par pays

    Le SRS a une limite concrète ici : les valeurs manquantes, les patterns d'outliers et
    les catégories d'articles **ne sont pas uniformément distribués entre pays**
    (mécanisme MAR — section 3).

    Le UK concentrant 65 % des transactions, un SRS de 200 lignes lui alloue ~130 lignes,
    laissant ~10 lignes pour l'Allemagne ou la France. Avec si peu d'observations par pays,
    les statistiques de nettoyage par strate (médiane d'imputation, seuils IQR) sont
    **instables ou absentes** pour les marchés minoritaires.

    En stratifiant par pays, chaque marché est représenté avec suffisamment d'observations
    pour calibrer correctement le pipeline sur **tous** les profils de qualité de données.
    """)

    _warning = mo.callout(
        mo.md("**Piège fréquent :** Stratifier sur une variable avec trop de modalités (ex. code postal) génère des strates de taille 1 ou 2 — le tirage devient déterministe. Regrouper les modalités rares avant de stratifier."),
        kind="warn",
    )
    _output = mo.vstack([_title, _principe, _example, _avantages])
    _output
    return


@app.cell
def _stratified_demo(retail):
    from sklearn.model_selection import train_test_split as _tts_strat

    _top5 = retail["Country"].value_counts().head(5).index
    _retail5 = retail[retail["Country"].isin(_top5)].copy()

    _strat_sample, _ = _tts_strat(
        _retail5, train_size=200, stratify=_retail5["Country"], random_state=42,
    )
    _srs_sample = _retail5.sample(n=200, random_state=123)

    _prop_pop   = _retail5["Country"].value_counts(normalize=True).sort_index()
    _prop_srs   = _srs_sample["Country"].value_counts(normalize=True).sort_index()
    _prop_strat = _strat_sample["Country"].value_counts(normalize=True).sort_index()

    comparison = pd.DataFrame({
        "Population (vraie)": _prop_pop,
        "SRS (n=200)": _prop_srs,
        "Stratifié (n=200)": _prop_strat,
    }).round(3)
    comparison.index.name = "Pays"

    print("=== Proportions par pays — impact sur la collecte de l'échantillon de développement ===")
    print(comparison)
    print()
    print("  → Le stratifié garantit la présence de chaque pays dans l'échantillon.")
    print("    Le pipeline sera exposé à tous les profils de qualité de données.")

    # Vérification clé : le taux de manquants CustomerID varie-t-il par pays ?
    print("\n=== Taux de manquants CustomerID par pays (défaut MAR — section 3) ===")
    _manq = _retail5.groupby("Country")["CustomerID"].apply(lambda x: x.isna().mean())
    _manq_srs = _srs_sample.groupby("Country")["CustomerID"].apply(lambda x: x.isna().mean())
    _manq_str = _strat_sample.groupby("Country")["CustomerID"].apply(lambda x: x.isna().mean())
    _taux = pd.DataFrame({
        "Pop. réelle": _manq,
        "SRS": _manq_srs,
        "Stratifié": _manq_str,
    }).round(3)
    print(_taux.to_string())
    print()
    print("  → Si ces taux varient entre pays, un SRS UK-dominant calibrera l'imputation")
    print("    sur le profil UK uniquement. Le stratifié révèle la diversité des mécanismes.")
    return


@app.cell
def _stratified_plot(retail):
    from sklearn.model_selection import train_test_split as _tts_sp

    _top5_sp = retail["Country"].value_counts().head(5).index
    _retail5_sp = retail[retail["Country"].isin(_top5_sp)].copy()
    _strat_s, _ = _tts_sp(_retail5_sp, train_size=200,
                           stratify=_retail5_sp["Country"], random_state=42)
    _srs_s = _retail5_sp.sample(n=200, random_state=123)

    _fig_sp, _axes_sp = plt.subplots(1, 3, figsize=(16, 4))

    for _ax_sp, _data_sp, _title_sp in zip(
        _axes_sp,
        [_retail5_sp, _srs_s, _strat_s],
        [f"Population (N={len(_retail5_sp):,})", "SRS (n=200)", "Stratifié (n=200)"],
    ):
        _vc = _data_sp["Country"].value_counts()
        _ax_sp.barh(_vc.index, _vc.values / len(_data_sp), color="#3498db", alpha=0.8, edgecolor="white")
        _ax_sp.set_title(_title_sp, fontweight="bold")
        _ax_sp.set_xlabel("Proportion")
        for _i_sp, _v_sp in enumerate(_vc.values / len(_data_sp)):
            _ax_sp.text(_v_sp + 0.002, _i_sp, f"{_v_sp:.1%}", va="center", fontsize=9)

    plt.suptitle("Comparaison : Population vs SRS vs Stratifié par pays", fontsize=12, y=1.02)
    plt.tight_layout()
    _fig_sp
    return


@app.cell(hide_code=True)
def _cluster_systemic_theory(mo):
    mo.md(rf"""
    ## 10.4 Grappes et systématique — échantillonner quand les données arrivent par lots
    En production, les données brutes n'arrivent pas en un seul bloc. Le dataset Online Retail
    illustre deux situations typiques de **collecte contrainte** :

    - Les transactions sont **partitionnées par pays** dans la base de données
    - Les exports arrivent sous forme de **fichiers CSV mensuels** triés chronologiquement
    Dans ces cas, le SRS et le stratifié ne sont pas directement applicables.

    ### 10.4.1 Échantillonnage par grappes — données partitionnées par pays
    On tire aléatoirement **quelques pays entiers** et on prend toutes leurs transactions.
    On accède ainsi uniquement aux partitions nécessaires, sans charger la table entière.

    {mo.image(src="images/cluster_sample.png", width=300)}

    | | Stratifié | Grappes |
    |--|-----------|---------|
    | Accès aux données | Toutes les strates | Quelques partitions seulement |
    | Interne à la grappe | Homogène | Hétérogène |
    | Objectif pipeline | Couvrir tous les profils de défauts | Réduire le coût de collecte |
    | Risque | Aucun si bien fait | Défauts d'un marché non couverts si mal tiré |

    ### 10.4.2 Échantillonnage systématique — exports CSV triés chronologiquement
    On lit le fichier séquentiellement et on retient **une ligne sur $k$** :

    $$k = \left\lfloor \frac{{N}}{{n}} \right\rfloor, \quad r \sim \mathcal{{U}}\{{1, \ldots, k\}} \qquad \text{{indices sélectionnés : }} r,\ r+k,\ r+2k,\ \ldots$$

    {mo.image(src="images/systematic_sample.png", width=300)}

    Dans l'exemple ci-dessus le pas est de 3, c-à-d  $k=3$.

    ⚠️ **Risque spécifique au pipeline** : si l'export est trié par montant de commande
    (grossistes B2B en tête, particuliers en queue), les règles d'outliers et les valeurs
    d'imputation calibrées sur cet échantillon ne seront pas transposables à la population.
    Toujours inspecter l'ordre du fichier source avant d'appliquer cette méthode.
    """)
    return


@app.cell
def _systematic_cluster_demo(retail):
    _mu = retail["UnitPrice"].mean()

    # === Échantillonnage systématique — simulation d'un export CSV ordonné ===
    print("=== Échantillonnage systématique (export CSV trié par date) ===")
    _N = len(retail)
    _n_target = 100
    _k = _N // _n_target
    _r = np.random.RandomState(42).randint(0, _k)
    _indices_syst = np.arange(_r, _N, _k)[:_n_target]
    _syst_sample = retail.iloc[_indices_syst]
    _err_syst = abs(_syst_sample['UnitPrice'].mean() - _mu) / _mu * 100
    print(f" Taille de la population: N={_N}")
    print(f"  pas k={_k}, départ aléatoire r={_r}")
    print(f"  UnitPrice estimé : {_syst_sample['UnitPrice'].mean():.2f} £  "
          f"(vrai : {_mu:.2f} £,  erreur : {_err_syst:.1f}%)")
    print(f"  Pays représentés : {_syst_sample['Country'].nunique()} sur {retail['Country'].nunique()}")

    # === Échantillonnage par grappes — base partitionnée par pays ===
    print("\n=== Échantillonnage par grappes ou cluster (base partitionnée par pays) ===")
    print("  Taille des grappes disponibles :")
    print(retail["Country"].value_counts().head(7).to_string())

    _grappes_choisies = np.random.RandomState(42).choice(
        retail["Country"].unique(), size=3, replace=False
    )
    _cluster_sample = retail[retail["Country"].isin(_grappes_choisies)]
    _err_clust = abs(_cluster_sample['UnitPrice'].mean() - _mu) / _mu * 100
    print(f"\n  Grappes tirées : {list(_grappes_choisies)}")
    print(f"  Transactions récupérées : {len(_cluster_sample):,}")
    print(f"  UnitPrice estimé : {_cluster_sample['UnitPrice'].mean():.2f} £  "
          f"(vrai : {_mu:.2f} £,  erreur : {_err_clust:.1f}%)")
    print(f"\n→ Les grappes peuvent être très inégales (UK domine). UK peut ne pas être tiré")
    print(f"  L'erreur dépend fortement des pays tirés au sort.")
    return


@app.cell(hide_code=True)
def _biais_theory(mo):
    mo.md(r"""
    ## 10.5 Biais d'échantillonnage et représentativité

    ### Définition

    Un **biais d'échantillonnage** est une erreur **systématique** dans la sélection des données :
    certains profils sont surreprésentés, d'autres ignorés. Contrairement à l'erreur aléatoire
    (qui diminue avec $n$), le biais **ne se corrige pas en augmentant la taille de l'échantillon**.
    L'échantillon n'est alors **pas représentatif** de la population.

    Dans un pipeline de nettoyage et de préparation, un biais de collecte se propage à toutes
    les étapes suivantes : les valeurs imputées, les seuils d'outliers, les encodages et les
    scalers calibrés sur un échantillon biaisé produiront des règles inadaptées à la population réelle.

    ### Biais typiques dans le pipeline Online Retail

    | Biais de collecte | Ce qui est mal capturé | Conséquence sur le nettoyage |
    |-------------------|----------------------|------------------------------|
    | **UK uniquement** (65 % du dataset) | Profils de défauts des autres marchés | Imputation et encodage UK-centrés |
    | **Grandes quantités > 50** (grossistes B2B) | Comportement des particuliers | Seuils IQR trop larges ; retours ignorés |
    | **Décembre uniquement** (Noël) | Saisonnalité hors fêtes | Valeurs de feature engineering faussées |
    | **Clients identifiés seulement** | 20 % de *guest checkout* | Taux de manquants CustomerID sous-estimé |

    > **Leçon fondamentale** : un grand $n$ ne corrige pas un biais de sélection.
    > Le *Literary Digest* (1936) s'est trompé sur la présidentielle américaine avec
    > **2,4 millions** de répondants — parce que son échantillon surreprésentait les ménages aisés.
    > La **représentativité prime toujours sur la taille**.
    """)
    return


@app.cell
def _biais_demo(retail):
    _vrai = retail["UnitPrice"].mean()

    # Cinq scénarios de sélection — du bon au pire
    _scenarios = {
        "✅  SRS (n=200) — référence": retail.sample(n=200, random_state=420)["UnitPrice"].mean(),
        "⚠️  UK uniquement (65% des lignes)": retail[retail["Country"]=="United Kingdom"]["UnitPrice"].mean(),
        "⚠️  Grandes quantités > 50 unités": retail[retail["Quantity"] > 50]["UnitPrice"].mean(),
        "⚠️  Clients identifiés seulement": retail[retail["CustomerID"].notna()]["UnitPrice"].mean(),
        "⚠️  Décembre uniquement (fêtes)": retail[retail["InvoiceDate"].dt.month == 12]["UnitPrice"].mean(),
    }

    _fig_b, _axes_b = plt.subplots(1, 2, figsize=(15, 5))

    # Panneau gauche : prix réel par pays (contexte du biais géographique)
    _top6 = retail["Country"].value_counts().head(6).index
    _prix_pays = retail[retail["Country"].isin(_top6)].groupby("Country")["UnitPrice"].mean().sort_values()
    _colors_pays = ["#e74c3c" if c == "United Kingdom" else "#3498db" for c in _prix_pays.index]
    _axes_b[0].barh(_prix_pays.index, _prix_pays.values, color=_colors_pays, alpha=0.8, edgecolor="white")
    _axes_b[0].axvline(_vrai, color="red", ls="--", lw=2, label=f"Moyenne globale = {_vrai:.2f} £")
    _axes_b[0].set_title("Prix moyen réel par pays\n(UK en rouge — marché dominant)",
                          fontweight="bold")
    _axes_b[0].set_xlabel("UnitPrice moyen (£)")
    _axes_b[0].legend(fontsize=9)
    for _i, _v in enumerate(_prix_pays.values):
        _axes_b[0].text(_v + 0.05, _i, f"{_v:.2f} £", va="center", fontsize=9)

    # Panneau droit : déformation de l'estimation selon le biais de sélection
    _vals = list(_scenarios.values())
    _labels = list(_scenarios.keys())
    _colors_sc = ["#2ecc71"] + ["#e74c3c"] * (len(_scenarios) - 1)
    _bars_b = _axes_b[1].barh(_labels, _vals, color=_colors_sc, alpha=0.85, edgecolor="white")
    _axes_b[1].axvline(_vrai, color="red", ls="--", lw=2, label=f"Vraie valeur = {_vrai:.2f} £")
    _axes_b[1].set_title("Déformation de l'estimation selon\nle mode de sélection des données",
                          fontweight="bold")
    _axes_b[1].set_xlabel("UnitPrice moyen estimé (£)")
    _axes_b[1].legend(fontsize=9)
    for _i, _v in enumerate(_vals):
        _ecart = (_v - _vrai) / _vrai * 100
        _txt = f"{_v:.2f} £  ({_ecart:+.1f}%)"
        _axes_b[1].text(_v + 0.05, _i, _txt, va="center", fontsize=8.5)

    plt.suptitle("Biais d'échantillonnage — Online Retail\n"
                 "Le SRS est sans biais (converge vers la vraie valeur en espérance) ;\n"
                 "les autres méthodes introduisent un biais structurel",
                 fontsize=11, fontweight="bold")
    plt.tight_layout()
    _fig_b
    return


@app.cell(hide_code=True)
def _bootstrap_theory(mo):
    mo.md(r"""
    ## 10.6 Bootstrap — valider la stabilité des statistiques calculées après nettoyage

    Après avoir imputé les valeurs manquantes et écrêté les outliers (sections 3 et 4),
    une question légitime se pose : **les statistiques calculées sur l'échantillon nettoyé
    sont-elles stables, ou très sensibles au tirage ?**

    La formule classique $\text{IC}_{95\%} = \bar{x} \pm 1{,}96 \cdot \frac{s}{\sqrt{n}}$
    suppose une distribution normale. Or le `UnitPrice` de l'Online Retail est fortement
    asymétrique (beaucoup de petits articles, quelques produits à 500 £) — la normalité
    n'est pas vérifiée. Le **Bootstrap** (Efron, 1979) contourne cette hypothèse.

    ### Principe — rééchantillonner avec remise depuis l'échantillon observé

    On simule ce que donneraient d'autres tirages depuis la même population, en
    **tirant avec remise** $B = 5\,000$ fois depuis l'échantillon observé :

    Pour chaque itération $b = 1, \ldots, B$ :

    1. Tirer $X^{*(b)}$ de taille $n$ **avec remise** dans l'échantillon nettoyé
    2. Recalculer la statistique $\hat{\theta}^{*(b)}$ (ex. médiane, seuil IQR)

    L'**IC à 95 %** est donné par les percentiles de la distribution bootstrap :

    $$\text{IC}_{95\%} = \left[Q_{2{,}5\%}(\hat{\theta}^*),\; Q_{97{,}5\%}(\hat{\theta}^*)\right]$$

    ### Usages concrets dans un pipeline de préparation

    - Vérifier que la **médiane imputée** est stable (IC étroit = imputation fiable)
    - Valider que les **bornes IQR** pour l'écrêtage des outliers ne varient pas trop d'un tirage à l'autre
    - Comparer les IC **avant et après nettoyage** pour détecter si une étape a introduit un biais
    """)
    return


@app.cell
def _bootstrap_demo(retail):
    # Bootstrap sur le UnitPrice moyen et médian — Online Retail
    _prices_bt = retail["UnitPrice"].dropna().values
    _prices_bt = _prices_bt[_prices_bt < np.percentile(_prices_bt, 99)]  # retire les outliers extrêmes
    _B = 5000
    _rng_bt = np.random.default_rng(42)
    _boot_means = np.empty(_B)
    _boot_medians = np.empty(_B)

    for _b in range(_B):
        _s = _rng_bt.choice(_prices_bt, size=len(_prices_bt), replace=True)
        _boot_means[_b] = _s.mean()
        _boot_medians[_b] = np.median(_s)

    _ic_moy = np.percentile(_boot_means, [2.5, 97.5])
    _ic_med = np.percentile(_boot_medians, [2.5, 97.5])

    print("=== BOOTSTRAP (B = 5 000) — UnitPrice — Online Retail ===")
    print()
    print(f"  Prix moyen   : {_prices_bt.mean():.3f} £  | IC95% bootstrap = [{_ic_moy[0]:.3f}, {_ic_moy[1]:.3f}]")
    print(f"  Prix médian  : {np.median(_prices_bt):.3f} £  | IC95% bootstrap = [{_ic_med[0]:.3f}, {_ic_med[1]:.3f}]")
    print()
    print(f"  Erreur-type bootstrap (moyenne) : {_boot_means.std(ddof=1):.4f} £")
    print(f"  Erreur-type théorique (σ/√n)   : {_prices_bt.std(ddof=1)/np.sqrt(len(_prices_bt)):.4f} £")
    print()
    print(f"  ► La médiane ({np.median(_prices_bt):.2f} £) << moyenne ({_prices_bt.mean():.2f} £)")
    print(f"    → distribution très asymétrique : quelques articles à prix élevé tirent la moyenne vers le haut.")
    print(f"    → pour le rapport au directeur, la médiane est plus représentative du 'prix typique'.")
    print(f"    → le bootstrap donne un IC sans hypothèse de normalité — indispensable ici.")

    _fig_bt, _axes_bt = plt.subplots(1, 2, figsize=(13, 4))

    _axes_bt[0].hist(_boot_means, bins=50, color="steelblue", edgecolor="white", alpha=0.7)
    _axes_bt[0].axvline(_prices_bt.mean(), color="red", lw=2, label=f"Estimation = {_prices_bt.mean():.3f} £")
    _axes_bt[0].axvline(_ic_moy[0], color="green", ls="--", label=f"IC95% = [{_ic_moy[0]:.3f}, {_ic_moy[1]:.3f}]")
    _axes_bt[0].axvline(_ic_moy[1], color="green", ls="--")
    _axes_bt[0].set_title("Distribution bootstrap — Moyenne du UnitPrice")
    _axes_bt[0].set_xlabel("UnitPrice moyen (£)")
    _axes_bt[0].legend(fontsize=9)

    _axes_bt[1].hist(_boot_medians, bins=50, color="coral", edgecolor="white", alpha=0.7)
    _axes_bt[1].axvline(np.median(_prices_bt), color="red", lw=2,
                        label=f"Estimation = {np.median(_prices_bt):.3f} £")
    _axes_bt[1].axvline(_ic_med[0], color="green", ls="--",
                        label=f"IC95% = [{_ic_med[0]:.3f}, {_ic_med[1]:.3f}]")
    _axes_bt[1].axvline(_ic_med[1], color="green", ls="--")
    _axes_bt[1].set_title("Distribution bootstrap — Médiane du UnitPrice")
    _axes_bt[1].set_xlabel("UnitPrice médian (£)")
    _axes_bt[1].legend(fontsize=9)

    plt.suptitle("Bootstrap — IC95% sur le prix unitaire (Online Retail)", fontsize=12, fontweight="bold")
    plt.tight_layout()
    _fig_bt
    return


@app.cell(hide_code=True)
def _classes_desequilibrees_theory(mo):
    mo.md(r"""
    ## 10.7 Classes déséquilibrées — dernière étape de préparation avant la modélisation

    Le pipeline des sections 1–7 produit un dataset nettoyé (voir TP section 12).
    Avant d'entraîner tout modèle de classification sur ces données, une dernière
    vérification s'impose : **la distribution des classes est-elle équilibrée ?**

    Dans le dataset Online Retail nettoyé, si l'on cherche à prédire les commandes
    à haute valeur (top 5 % du revenu), on obtient une cible très déséquilibrée :

    ```
    Commandes normales    : 95 %
    Commandes haute valeur :  5 %
    ```

    Un modèle naïf qui prédit toujours "normal" atteint **95 % d'accuracy**
    sans jamais détecter une seule commande haute valeur — c'est le **piège de l'accuracy**
    sur classes déséquilibrées.

    Ce déséquilibre est un **problème de préparation des données**, pas un problème
    de choix d'algorithme. Il doit être traité avant la modélisation.

    ### Trois familles de solutions — dans l'ordre de préférence pratique

    1. **Modifier la préparation des données** (au niveau de l'échantillon d'entraînement) :
       - **Undersampling** : retirer des exemples de la classe majoritaire — rapide, perd de l'info
       - **SMOTE** : générer synthétiquement de nouveaux exemples minoritaires — préserve l'information
    2. **Modifier l'algorithme** : `class_weight='balanced'` — aucune modification des données,
       mais tous les algorithmes ne le supportent pas
    3. **Modifier l'évaluation** : toujours mesurer **F1, AUC-ROC, rappel** — jamais l'accuracy seule

    ### SMOTE — Synthetic Minority Over-sampling Technique (Chawla et al., 2002)

    SMOTE génère des exemples synthétiques par **interpolation** entre un exemple
    minoritaire $x_i$ et l'un de ses $k$ plus proches voisins dans la classe minoritaire :

    $$x_{\text{new}} = x_i + \lambda \cdot (x_{nn} - x_i), \quad \lambda \sim \mathcal{U}(0,1)$$

    > ⚠️ **Règle critique de pipeline** : SMOTE s'applique **uniquement sur le train set**,
    > après le split train/test. L'appliquer avant revient à faire fuiter des exemples synthétiques
    > dans le test set — les performances mesurées seront artificiellement gonflées (data leakage).
    """)
    return


@app.cell
def _smote_demo(retail):
    from sklearn.preprocessing import StandardScaler as _SS_sm
    from imblearn.over_sampling import SMOTE as _SMOTE
    from imblearn.under_sampling import RandomUnderSampler as _RUS
    from sklearn.linear_model import LogisticRegression as _LR
    from sklearn.metrics import classification_report as _cr, roc_auc_score as _auc

    # Cible : transactions à haute valeur (top 5% du Revenue) — classe minoritaire réaliste
    _df_sm = retail[["UnitPrice", "Quantity"]].copy()
    _df_sm = _df_sm[(_df_sm["UnitPrice"] > 0) & (_df_sm["Quantity"] > 0)].reset_index(drop=True)
    _rev_sm = _df_sm["UnitPrice"] * _df_sm["Quantity"]
    _seuil = _rev_sm.quantile(0.95)
    _y_sm_raw = (_rev_sm >= _seuil).astype(int)

    # Limiter les outliers pour la visualisation
    _p99u = _df_sm["UnitPrice"].quantile(0.99)
    _p99q = _df_sm["Quantity"].quantile(0.99)
    _mask_vis = (_df_sm["UnitPrice"] <= _p99u) & (_df_sm["Quantity"] <= _p99q)
    _X_vis = _df_sm[_mask_vis][["UnitPrice", "Quantity"]].values
    _y_vis = _y_sm_raw[_mask_vis].values

    print("=== ONLINE RETAIL — Prédire les commandes haute valeur (top 5% Revenue) ===")
    print()
    print(f"  Seuil 'haute valeur' : Revenue ≥ {_seuil:.1f} £·unités")
    print(f"  Commandes normales   : {(_y_vis==0).sum():,}  ({1-_y_vis.mean():.1%})")
    print(f"  Commandes HV         : {(_y_vis==1).sum():,}  ({_y_vis.mean():.1%})  ← classe minoritaire")
    print()
    print("  → Un modèle qui prédit toujours 'normal' aurait 95% d'accuracy.")
    print("    Il raterait 100% des commandes haute valeur. C'est le piège de l'accuracy.")
    print()

    _scaler_sm = _SS_sm()
    _X_vis_s = _scaler_sm.fit_transform(_X_vis)

    _X_smote, _y_smote = _SMOTE(random_state=42, k_neighbors=5).fit_resample(_X_vis_s, _y_vis)
    _X_under, _y_under = _RUS(random_state=42).fit_resample(_X_vis_s, _y_vis)
    print(f"  Après SMOTE         : {np.bincount(_y_smote)}  (synthèse de {(_y_smote==1).sum()-(_y_vis==1).sum()} exemples HV)")
    print(f"  Après undersampling : {np.bincount(_y_under)}  (retrait de {(_y_vis==0).sum()-(_y_under==0).sum()} commandes normales)")

    # Visualisation dans l'espace standardisé
    _fig_sm, _axes_sm = plt.subplots(1, 3, figsize=(16, 5))
    _datasets_sm = [
        (_X_vis_s, _y_vis, f"Original déséquilibré\n(95% normal / 5% haute valeur)"),
        (_X_smote, _y_smote, "Après SMOTE\n(synthèse de la classe minoritaire)"),
        (_X_under, _y_under, "Après Undersampling\n(réduction de la classe majoritaire)"),
    ]
    for _ax_sm, (_Xv, _yv, _t) in zip(_axes_sm, _datasets_sm):
        _ax_sm.scatter(_Xv[_yv==0, 0], _Xv[_yv==0, 1], s=8, alpha=0.3, c="steelblue",
                       label=f"Normal (n={(_yv==0).sum():,})")
        _ax_sm.scatter(_Xv[_yv==1, 0], _Xv[_yv==1, 1], s=15, alpha=0.7, c="crimson",
                       label=f"Haute valeur (n={(_yv==1).sum():,})")
        _ax_sm.set_title(_t, fontweight="bold")
        _ax_sm.set_xlabel("UnitPrice (standardisé)")
        _ax_sm.set_ylabel("Quantity (standardisé)")
        _ax_sm.legend(fontsize=8)

    plt.suptitle("Online Retail — SMOTE vs Undersampling sur la détection de transactions haute valeur",
                 fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _fig_sm
    return


@app.cell(hide_code=True)
def _sampling_conclusion(mo):
    mo.md(r"""
    ---
    ### 💡 Récapitulatif — L'échantillonnage dans le pipeline de données

    Chaque technique répond à une contrainte ou un problème précis qui surgit
    à une étape du pipeline **Collecte → Nettoyage → Préparation → Modélisation**.

    | Étape du pipeline | Contrainte / Problème | Technique |
    |-------------------|-----------------------|-----------|
    | **Collecte** — dataset trop volumineux | Impossible de traiter toute la population | **SRS** (SRSWOR) — sous-ensemble représentatif pour tester le pipeline |
    | **Collecte** — données partitionnées | Accès limité : export par pays ou par fichier CSV | **Grappes** (pays entiers) ou **Systématique** (1 ligne sur $k$) |
    | **Audit qualité** — manquants non uniformes (MAR) | `CustomerID` absent davantage hors UK → SRS manque les marchés minoritaires | **Stratifié** par pays — chaque strate calibre ses propres règles de nettoyage |
    | **Nettoyage** — statistiques calibrées sur données biaisées | Imputation médiane, IQR calculés sur un échantillon B2B ou saisonnier | **Analyse des biais** — la représentativité prime sur la taille $n$ |
    | **Validation des statistiques nettoyées** — prix asymétriques | IC classique (normalité) invalide pour `UnitPrice` → IC instable | **Bootstrap** — IC percentile sans hypothèse, valide sur distributions quelconques |
    | **Préparation avant modélisation** — cible déséquilibrée | Top 5 % haute valeur : accuracy trompeuse (95 % sans rien détecter) | **SMOTE** sur train uniquement → évaluer avec F1 / AUC, pas l'accuracy |

    > Le fil conducteur : **on ne choisit pas une technique par curiosité, on la choisit
    > parce qu'une étape précédente du pipeline l'exige.**
    """)
    return


@app.cell(hide_code=True)
def _section_tp(mo):
    mo.md(r"""
    ---
    ## 12. TP — Pipeline de nettoyage complet

    ### Objectif

    Construire un pipeline **reproductible** en 5 étapes, applicable à tout dataset e-commerce similaire à Online Retail (UCI ML Repository).

    ```
    ┌─────────────────────────────────────────────────────┐
    │  PIPELINE — Online Retail Dataset                    │
    ├────────────┬────────────────────────────────────────┤
    │  Étape 1   │  Suppression des doublons              │
    │  Étape 2   │  Filtrage des lignes invalides         │
    │  Étape 3   │  Imputation des valeurs manquantes     │
    │  Étape 4   │  Traitement des outliers (capping)     │
    │  Étape 5   │  Encodage + Standardisation            │
    └────────────┴────────────────────────────────────────┘
    ```
    """)
    return


@app.cell
def _tp_pipeline(df_retail):
    from sklearn.preprocessing import StandardScaler as _SS12, LabelEncoder as _LE12
    from sklearn.impute import SimpleImputer as _SI12

    print("=" * 60)
    print("PIPELINE DE NETTOYAGE — Online Retail")
    print("=" * 60)

    _df12 = df_retail.copy()
    _n0 = len(_df12)
    print(f"\n📊 Données initiales : {_n0:,} lignes")

    # Étape 1 : Doublons
    _df12 = _df12.drop_duplicates()
    print(f"\n✅ Étape 1 — Doublons : -{_n0 - len(_df12)} lignes → {len(_df12):,} restantes")

    # Étape 2 : Lignes invalides
    _df_retours12 = _df12[_df12["Quantity"] < 0].copy()
    _df12 = _df12.dropna(subset=["StockCode"])
    _df12 = _df12[~_df12["Description"].isin(["???", "SAMPLE"])]
    _df12 = _df12[_df12["Quantity"] > 0]
    print(f"\n✅ Étape 2 — Invalides filtrés")
    print(f"   Retours isolés : {len(_df_retours12)}")
    print(f"   Restant : {len(_df12):,} lignes")

    # Étape 3 : Imputation
    _imp12 = _SI12(strategy="median")
    _df12["UnitPrice"] = _imp12.fit_transform(_df12[["UnitPrice"]]).flatten()
    _df12["HasCustomerID"] = _df12["CustomerID"].notna().astype(int)
    _df12["CustomerID"] = _df12["CustomerID"].fillna(-1)
    _country_mode12 = _df12["Country"].mode()[0]
    _df12["Country"] = _df12["Country"].fillna(_country_mode12)
    print(f"\n✅ Étape 3 — Imputation")
    print(f"   UnitPrice médiane = {_imp12.statistics_[0]:.4f} €")
    print(f"   CustomerID : indicatrice HasCustomerID créée")
    print(f"   Country manquants → '{_country_mode12}'")
    print(f"   Valeurs manquantes restantes : {_df12.isnull().sum().sum()}")

    # Étape 4 : Capping outliers (3×IQR)
    for _col12 in ["Quantity", "UnitPrice"]:
        _q1_12 = _df12[_col12].quantile(0.25)
        _q3_12 = _df12[_col12].quantile(0.75)
        _iqr12 = _q3_12 - _q1_12
        _cap12 = _q3_12 + 3 * _iqr12
        _out12 = (_df12[_col12] > _cap12).sum()
        _df12[f"{_col12}_capped"] = _df12[_col12].clip(upper=_cap12)
        print(f"\n✅ Étape 4 — Capping {_col12} : {_out12} outliers → cap={_cap12:.2f}")

    # Étape 5 : Encodage + standardisation
    _le12 = _LE12()
    _df12["Country_encoded"] = _le12.fit_transform(_df12["Country"])
    _ss12 = _SS12()
    _cols_num12 = ["Quantity_capped", "UnitPrice_capped"]
    _df12_scaled = _df12.copy()
    _df12_scaled[_cols_num12] = _ss12.fit_transform(_df12[_cols_num12])
    print(f"\n✅ Étape 5 — Encodage + standardisation")
    print(f"   Country : {_le12.classes_.__len__()} catégories encodées")
    print(f"   Colonnes standardisées : {_cols_num12}")

    # Revenue
    _df12["Revenue"] = _df12["Quantity_capped"] * _df12["UnitPrice_capped"]

    print(f"\n{'='*60}")
    print(f"RÉSUMÉ PIPELINE")
    print(f"{'='*60}")
    print(f"  Entrée  : {_n0:,} lignes")
    print(f"  Sortie  : {len(_df12):,} lignes ({len(_df12)/_n0*100:.1f}% conservées)")
    print(f"  CA total : {_df12['Revenue'].sum():,.0f} £×unités")
    print(f"  ✅ Dataset prêt pour EDA (S3), Clustering (S5), ML (S6)")

    df_pipe = _df12
    return (df_pipe,)


@app.cell
def _plot_pipeline(df_pipe, df_retail):
    _fig13, _axes13 = plt.subplots(2, 3, figsize=(16, 9))

    # Qty : avant vs après
    _ax13a = _axes13[0, 0]
    _qty_avant13 = df_retail["Quantity"].dropna()
    _qty_apres13 = df_pipe["Quantity_capped"]
    _p99_13 = min(_qty_avant13.quantile(0.99), 300)
    _bins13 = np.linspace(0, _p99_13, 40)
    _ax13a.hist(_qty_avant13[(_qty_avant13 >= 0) & (_qty_avant13 <= _p99_13)], bins=_bins13, alpha=0.5, color="#e74c3c", label="Avant")
    _ax13a.hist(_qty_apres13[_qty_apres13 <= _p99_13], bins=_bins13, alpha=0.5, color="#2ecc71", label="Après")
    _ax13a.set_title("Quantity — avant vs après", fontweight="bold")
    _ax13a.legend()

    # Price : avant vs après
    _ax13b = _axes13[0, 1]
    _pr_avant13 = df_retail["UnitPrice"].dropna()
    _pr_apres13 = df_pipe["UnitPrice_capped"]
    _p99p = min(_pr_avant13.quantile(0.99), 50)
    _bins13p = np.linspace(0, _p99p, 40)
    _ax13b.hist(_pr_avant13[_pr_avant13 <= _p99p], bins=_bins13p, alpha=0.5, color="#e74c3c", label="Avant")
    _ax13b.hist(_pr_apres13[_pr_apres13 <= _p99p], bins=_bins13p, alpha=0.5, color="#2ecc71", label="Après")
    _ax13b.set_title("UnitPrice — avant vs après", fontweight="bold")
    _ax13b.legend()

    # Taux de complétude
    _ax13c = _axes13[0, 2]
    _cols_com13 = [c for c in df_retail.columns if c in df_pipe.columns]
    _comp_av13 = (1 - df_retail[_cols_com13].isnull().mean()) * 100
    _comp_ap13 = (1 - df_pipe[_cols_com13].isnull().mean()) * 100
    _x13c = range(len(_cols_com13))
    _ax13c.bar([xi - 0.2 for xi in _x13c], _comp_av13, 0.4, alpha=0.7, color="#e74c3c", label="Avant")
    _ax13c.bar([xi + 0.2 for xi in _x13c], _comp_ap13, 0.4, alpha=0.7, color="#2ecc71", label="Après")
    _ax13c.set_xticks(list(_x13c))
    _ax13c.set_xticklabels(_cols_com13, rotation=45, ha="right", fontsize=8)
    _ax13c.set_title("Taux de complétude (%)", fontweight="bold")
    _ax13c.set_ylim(70, 105)
    _ax13c.legend()

    # Distribution Revenue
    _ax13d = _axes13[1, 0]
    _rev13 = df_pipe["Revenue"]
    _rev13_c = _rev13[_rev13 < _rev13.quantile(0.95)]
    _ax13d.hist(_rev13_c, bins=40, color="#3498db", alpha=0.8, edgecolor="white")
    _ax13d.axvline(_rev13_c.median(), color="red", linestyle="--", label=f"Médiane={_rev13_c.median():.1f}")
    _ax13d.set_title("Distribution du Revenue\n(après nettoyage)", fontweight="bold")
    _ax13d.legend()

    # Pays distribution après nettoyage
    _ax13e = _axes13[1, 1]
    _country13 = df_pipe["Country"].value_counts().head(6)
    _ax13e.barh(_country13.index, _country13.values, color=["#3498db","#2ecc71","#e74c3c","#f39c12","#9b59b6","#1abc9c"])
    _ax13e.set_title("Distribution pays\n(après nettoyage)", fontweight="bold")
    for _i13, _v13 in enumerate(_country13.values):
        _ax13e.text(_v13 + 0.5, _i13, str(_v13), va="center", fontsize=9)

    # Synthèse texte
    _ax13f = _axes13[1, 2]
    _ax13f.axis("off")
    _txt13 = (
        "BILAN DU PIPELINE\n"
        "─────────────────────\n"
        f"📥 Entrée  : {len(df_retail):,} lignes\n"
        f"📤 Sortie  : {len(df_pipe):,} lignes\n"
        f"💾 Rétention : {len(df_pipe)/len(df_retail)*100:.1f}%\n"
        "─────────────────────\n"
        "✅ Doublons supprimés\n"
        "✅ Invalides filtrés\n"
        "✅ UnitPrice → médiane\n"
        "✅ CustomerID → indicatrice\n"
        "✅ Outliers → capping 3×IQR\n"
        "✅ Encodage + StandardScaler\n"
        "─────────────────────\n"
        "→ Prêt pour EDA (S3)\n"
        "→ Prêt pour Clustering (S5)\n"
        "→ Prêt pour ML (S6)"
    )
    _ax13f.text(0.05, 0.95, _txt13, transform=_ax13f.transAxes, va="top",
                fontsize=10, fontfamily="monospace",
                bbox=dict(boxstyle="round", facecolor="#ecf0f1", alpha=0.8))

    plt.suptitle("Avant vs Après — Pipeline de nettoyage Online Retail", fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig("/tmp/pipeline_bilan.png", dpi=120, bbox_inches="tight")
    plt.show()
    return


@app.cell(hide_code=True)
def _synthese(mo):
    mo.md(r"""
    ---
    ## 🎯 Synthèse de la séance

    ### Récapitulatif

    | Concept | Méthode(s) clés | Outil Python |
    |---------|----------------|-------------|
    | Collecte | CSV, JSON, API REST, SQL | `pd.read_csv`, `requests`, `sqlalchemy` |
    | Audit qualité | Profiling, patterns manquants | `df.info()`, `missingno`, `ydata-profiling` |
    | Valeurs manquantes | MCAR/MAR/MNAR → stratégie | `SimpleImputer`, `KNNImputer`, `IterativeImputer` |
    | Outliers | IQR, Z-score, Isolation Forest | `scipy.stats`, `IsolationForest` |
    | Mise à l'échelle | Min-Max, Z-score, Robust | `MinMaxScaler`, `StandardScaler`, `RobustScaler` |
    | Encodage | Label, OneHot, Fréquence | `LabelEncoder`, `OneHotEncoder` |
    | Feature Engineering | Dates, binning, combinaison | `pd.cut()`, `pd.qcut()`, `.dt` |
    | Pipeline complet | Sous-pipelines num/cat | `ColumnTransformer`, `Pipeline` |
    | Échantillonnage | SRS, Stratifié, Grappes, Systématique | `df.sample()`, `train_test_split(stratify=)` |
    | Bootstrap | IC non-paramétrique, erreur-type | `np.random.default_rng`, percentiles |
    | Classes déséquilibrées | SMOTE, undersampling, class_weight | `imblearn.over_sampling.SMOTE` |

    ### Règles d'or

    1. **Diagnostiquer avant de traiter** — `df.info()`, `df.describe()`, `missingno`
    2. **Identifier le mécanisme** (MCAR/MAR/MNAR) avant de choisir la stratégie
    3. **Fitter uniquement sur le train set** — jamais sur le test set (data leakage)
    4. **Séparer num/cat dans un ColumnTransformer** — pipeline reproductible et extensible
    5. **Conserver les données brutes** — toujours travailler sur des copies
    6. **Documenter chaque choix** — pipeline sauvegardé avec `joblib.dump()`
    7. **Représentativité > taille** — un grand $n$ ne corrige pas un biais de sélection
    8. **SMOTE uniquement sur le train set** — jamais sur le test set
    9. **Accuracy ≠ performance** sur classes déséquilibrées — utiliser F1, AUC, rappel

    ### Prochaine séance (Séance 3)

    **Analyse statistique et visualisation exploratoire :**
    Distributions · Tests de normalité · Corrélations · Tests d'hypothèses · Visualisations seaborn/Plotly

    ---

    ### 📚 Références

    - Wes McKinney — *Python for Data Analysis*, O'Reilly, 3e éd. (2022) — **ch. 6 & 7**
    - Jake VanderPlas — *Python Data Science Handbook*, O'Reilly — **ch. 5 (Feature Engineering)**
    - Rubin, D.B. (1976) — *Inference and Missing Data*, Biometrika, 63(3)
    - scikit-learn — [Imputation de valeurs manquantes](https://scikit-learn.org/stable/modules/impute.html)
    - Online Retail Dataset — [UCI ML Repository](https://archive.ics.uci.edu/dataset/352/online+retail)

    ---
    """)
    return


@app.cell
def _():
    return


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
